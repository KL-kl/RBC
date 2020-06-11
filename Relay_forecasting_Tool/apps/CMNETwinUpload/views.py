import os
from django.http import HttpResponse, JsonResponse
from django.db import connection
from django.shortcuts import render, HttpResponse
from django.utils.encoding import escape_uri_path
import xlrd
import xlwt
from xlutils.copy import copy
import numpy as np
import re
from string import digits
import time
import random
from openpyxl import load_workbook
import logging
import json
from functools import reduce
from Relay_forecasting_Tool import settings
from CMNETwinUpload.models import ProjectInfo, CNodes_detail, CMetricTable
from utils.hebin import *
from utils.sort_path import *
from utils.fault_polling_min import *
from utils.fault_polling import *
from utils.daikuanfenpei import bandwidth_allocation
from utils.zhongjidaikuan100GE import *
from utils.zhongjidaikuan10GE import *
from utils.zhongjidaikuan400GE import *
from utils.excel_style import *
from utils.calculateTable import *

logger = logging.getLogger('log')

# 自定义表格样式
style = xlwt.easyxf('align: wrap on, vert centre, horiz center;'
                    'borders:left 1,right 1,top 1,bottom 1')
style1 = xlwt.easyxf('align: wrap on, vert centre, horiz center;'
                     'borders:left 1,right 1,top 1,bottom 1')
style2 = xlwt.easyxf('font: bold on,name Calibri;align: wrap on, vert centre, horiz center;'
                     'borders:left 1,right 1,top 1,bottom 1')

style3 = xlwt.easyxf('font: name Calibri;align: wrap on, vert centre, horiz center;'
                     'borders:left 1,right 1,top 1,bottom 1')

he_style3 = easyxf('pattern: pattern solid, fore_colour ice_blue; '
                   'font: bold on,name Calibri,height 260;'
                   'align: wrap on, vert centre, horiz center;'
                   'borders:left 2,right 2,top 2,bottom 2')
gz_style = xlwt.easyxf('pattern: pattern solid, fore_colour red;')


def CMNETwin(request):
    return render(request, 'CMNETwin.html')

def CMNETwin_2(request):
    return render(request, 'cmnet_win.html')

def show_progress1(request):
    return render(request, 'show.html')


def show_progress2(request):
    return render(request, '测试.html')


# 数据库查询节点信息，网络层次信息等
def find_node(pm):
    data1 = CNodes_detail.objects.filter(plane2=pm, part_subdivide2='BB-c').exclude(program='新建').values(
        'province', 'part_subdivide2',
        'devicename2')
    data2 = CNodes_detail.objects.filter(plane2=pm, part_subdivide2='BB-a').exclude(program='新建').values(
        'province', 'part_subdivide2',
        'devicename2')
    data3 = CNodes_detail.objects.filter(plane2=pm, part_subdivide2='BC').exclude(program='新建').values(
        'province', 'part_subdivide2',
        'devicename2')

    hx = {}
    lhx = {}
    hj = {}
    node_dict = {}
    for i in data1:
        if len(i['province']) > 2:
            nick = i['province'][:2]
        else:
            nick = i['province']

        res = i['devicename2'].split('_')  # ['北京', 'N2', 'BB', '5']        #
        nick_pro = nick + res[len(res) - 1]  # 北京5

        hx[nick_pro] = i['devicename2']
    for i in data2:
        if len(i['province']) > 2:
            nick = i['province'][:2]
        else:
            nick = i['province']

        res = i['devicename2'].split('_')  # ['北京', 'N2', 'BB', '5']        #
        nick_pro = nick + res[len(res) - 1]  # 北京5

        lhx[nick_pro] = i['devicename2']

    for i in data3:

        if len(i['province']) > 2:
            nick = i['province'][:2]
        else:
            nick = i['province']

        res = i['devicename2'].split('_')  # ['北京', 'N2', 'BB', '5']        #
        nick_pro = nick + res[len(res) - 1]  # 北京5

        hj[nick_pro] = i['devicename2']

    node_dict.update(hx)
    node_dict.update(lhx)
    node_dict.update(hj)

    return node_dict, hx, lhx, hj  # 节点字典


# 生成随机数
def randomnum():
    str = ""  # 随机数
    for i in range(3):
        ch = chr(random.randrange(ord('0'), ord('9') + 1))  # ord(char)函数将char类型的单字符转换成ASCII码值

        str += ch  # print(str) # 例如：202004281817029441

    str = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())) + str

    return str


# CMNET 节点上传，节点信息入库，创建metric及传输距离模板表
def upload_Cfile(request):
    if request.method == 'POST':
        pro_name = request.POST.get('pro_name')
        pm = request.POST.get('pm')
        myFile = request.FILES.get('node_file')

        if not all([pro_name, pm, myFile]):
            # 有数据为空
            return JsonResponse({'status': 'err', 'msg': '参数不能为空!'})

        path = settings.MEDIA_ROOT + 'cmnetwin/nodes/'  # 上传文件的保存路径，可以自己指定任意的路径

        # 创建模板
        model_path = os.path.abspath('models')
        if not os.path.exists(model_path):
            os.makedirs(model_path)

        if not os.path.exists(path):
            logger.info('没有存储文件路径，创建新文件')
            os.makedirs(path)

        with open(os.path.join(path, myFile.name), 'wb+') as f:
            for chunk in myFile.chunks():
                f.write(chunk)

        node_list = CNodes_detail.objects.all()
        if len(node_list) > 0:

            cursor = connection.cursor()  # 建立游标

            cursor.execute("TRUNCATE TABLE e_cnodes_detail")
            cursor.execute("ALTER TABLE e_cnodes_detail AUTO_INCREMENT = 1")

            cursor.close()  # 关闭游标

            excel = xlrd.open_workbook(path + myFile.name)
            sheet = excel.sheet_by_index(1)
            nrows = sheet.nrows
        else:
            cursor = connection.cursor()  # 建立游标
            cursor.execute("ALTER TABLE e_cnodes_detail AUTO_INCREMENT = 1")
            cursor.close()  # 关闭游标
            excel = xlrd.open_workbook(path + myFile.name)
            sheet = excel.sheet_by_index(1)
            nrows = sheet.nrows

        for i in range(3, nrows):
            cnode = CNodes_detail()
            cnode.province = sheet.cell(i, 1).value
            cnode.city = sheet.cell(i, 2).value
            cnode.office_address1 = sheet.cell(i, 3).value
            cnode.building_no1 = sheet.cell(i, 4).value
            cnode.floor1 = sheet.cell(i, 5).value
            cnode.room_num1 = sheet.cell(i, 6).value
            cnode.plane1 = sheet.cell(i, 7).value
            cnode.network_level1 = sheet.cell(i, 8).value
            cnode.part1 = sheet.cell(i, 9).value
            cnode.part_subdivide1 = sheet.cell(i, 10).value
            cnode.devicename1 = sheet.cell(i, 11).value
            cnode.device_comp1 = sheet.cell(i, 12).value
            cnode.unit_type1 = sheet.cell(i, 13).value
            cnode.device_state1 = sheet.cell(i, 14).value
            cnode.program = sheet.cell(i, 15).value
            cnode.office_address2 = sheet.cell(i, 16).value
            cnode.building_no2 = sheet.cell(i, 17).value
            cnode.floor2 = sheet.cell(i, 18).value
            cnode.room_num2 = sheet.cell(i, 19).value
            cnode.plane2 = sheet.cell(i, 20).value
            cnode.network_level2 = sheet.cell(i, 21).value
            cnode.part2 = sheet.cell(i, 22).value
            cnode.part_subdivide2 = sheet.cell(i, 23).value
            cnode.devicename2 = sheet.cell(i, 24).value
            cnode.device_comp2 = sheet.cell(i, 25).value
            cnode.unit_type2 = sheet.cell(i, 26).value
            cnode.device_state2 = sheet.cell(i, 27).value
            cnode.save()

        pro = ProjectInfo()
        pro.id = 2
        pro.name = pro_name
        pro.pro_type = 2
        pro.pm = pm
        pro.nodefile = os.path.join(path, myFile.name)
        pro.createBy_id = request.user.id
        pro.save()

        if pm == 'N2':
            node_dict, hx, lhx, hj = find_node(pm)
            ll = list(node_dict.keys())

            work_book1 = xlwt.Workbook(encoding='utf-8')
            sheet1 = add_sheet(work_book1,'Metric')
            sheet1.write(0, 0, 'Metric')

            work_book2 = xlwt.Workbook(encoding='utf-8')
            sheet2 = add_sheet(work_book2,'传输距离')
            sheet2.write(0, 0, '传输距离')
            for i in range(len(ll)):
                sheet1.write(0, i + 1, ll[i])
                sheet1.write(i + 1, 0, ll[i])
                sheet2.write(0, i + 1, ll[i])
                sheet2.write(i + 1, 0, ll[i])
            work_book1.save(os.path.join(os.path.abspath('models'), 'CMNET网_metric.xls'))
            work_book2.save(os.path.join(os.path.abspath('models'), 'CMNET网_传输距离.xls'))
        elif pm == 'N1':
            node_dict, hx, lhx, hj = find_node(pm)
            ll = list(node_dict.keys())
            work_book1 = xlwt.Workbook(encoding='utf-8')
            sheet1 = add_sheet(work_book1,'Metric')
            sheet1.write(0, 0, 'Metric')

            work_book2 = xlwt.Workbook(encoding='utf-8')
            sheet2 = add_sheet(work_book2,'传输距离')
            sheet2.write(0, 0, '传输距离')
            for i in range(len(ll)):
                sheet1.write(0, i + 1, ll[i])
                sheet1.write(i + 1, 0, ll[i])
                sheet2.write(0, i + 1, ll[i])
                sheet2.write(i + 1, 0, ll[i])
            work_book1.save(os.path.join(model_path, 'CMNET网_metric.xls'))
            work_book2.save(os.path.join(model_path, 'CMNET网_传输距离.xls'))

        logger.info('上传成功！！！')
        return JsonResponse({'status': 'ok', 'msg': '上传成功！！！'})


# 更新节点信息
def upload_newCfile(request):
    if request.method == 'POST':
        myFile = request.FILES.get('node_file')
        if not myFile:
            # 有数据为空
            return JsonResponse({'status': 'err', 'msg': '参数不能为空!'})

        path = settings.MEDIA_ROOT + 'cmnetwin/nodes/'  # 上传文件的保存路径，可以自己指定任意的路径

        # 创建模板
        model_path = os.path.abspath('models')
        if not os.path.exists(model_path):
            os.makedirs(model_path)

        if not os.path.exists(path):
            logger.info('没有存储文件路径，创建新文件')
            os.makedirs(path)

        with open(os.path.join(path, myFile.name), 'wb+') as f:
            for chunk in myFile.chunks():
                f.write(chunk)

        node_list = CNodes_detail.objects.all()
        if len(node_list) > 0:

            cursor = connection.cursor()  # 建立游标

            cursor.execute("TRUNCATE TABLE e_cnodes_detail")
            cursor.execute("ALTER TABLE e_cnodes_detail AUTO_INCREMENT = 1")

            cursor.close()  # 关闭游标

            excel = xlrd.open_workbook(path + myFile.name)
            sheet = excel.sheet_by_index(1)
            nrows = sheet.nrows
        else:
            cursor = connection.cursor()  # 建立游标
            cursor.execute("ALTER TABLE e_cnodes_detail AUTO_INCREMENT = 1")
            cursor.close()  # 关闭游标
            excel = xlrd.open_workbook(path + myFile.name)
            sheet = excel.sheet_by_index(1)
            nrows = sheet.nrows

        for i in range(3, nrows):
            cnode = CNodes_detail()
            cnode.province = sheet.cell(i, 1).value
            cnode.city = sheet.cell(i, 2).value
            cnode.office_address1 = sheet.cell(i, 3).value
            cnode.building_no1 = sheet.cell(i, 4).value
            cnode.floor1 = sheet.cell(i, 5).value
            cnode.room_num1 = sheet.cell(i, 6).value
            cnode.plane1 = sheet.cell(i, 7).value
            cnode.network_level1 = sheet.cell(i, 8).value
            cnode.part1 = sheet.cell(i, 9).value
            cnode.part_subdivide1 = sheet.cell(i, 10).value
            cnode.devicename1 = sheet.cell(i, 11).value
            cnode.device_comp1 = sheet.cell(i, 12).value
            cnode.unit_type1 = sheet.cell(i, 13).value
            cnode.device_state1 = sheet.cell(i, 14).value
            cnode.program = sheet.cell(i, 15).value
            cnode.office_address2 = sheet.cell(i, 16).value
            cnode.building_no2 = sheet.cell(i, 17).value
            cnode.floor2 = sheet.cell(i, 18).value
            cnode.room_num2 = sheet.cell(i, 19).value
            cnode.plane2 = sheet.cell(i, 20).value
            cnode.network_level2 = sheet.cell(i, 21).value
            cnode.part2 = sheet.cell(i, 22).value
            cnode.part_subdivide2 = sheet.cell(i, 23).value
            cnode.devicename2 = sheet.cell(i, 24).value
            cnode.device_comp2 = sheet.cell(i, 25).value
            cnode.unit_type2 = sheet.cell(i, 26).value
            cnode.device_state2 = sheet.cell(i, 27).value
            cnode.save()

        # 更新数据库工程信息
        ProjectInfo.objects.filter(createBy=request.user.id, id=2).update(
            nodefile=os.path.join(path, myFile.name), pro_type=2)

        # 查询CMNET网平面
        pm = ProjectInfo.objects.filter(createBy=request.user.id, pro_type=2).values('pm')

        if pm == 'N2':
            node_dict, hx, lhx, hj = find_node(pm)
            ll = list(node_dict.keys())

            work_book1 = xlwt.Workbook(encoding='utf-8')
            sheet1 = add_sheet(work_book1,'Metric')
            sheet1.write(0, 0, 'Metric')

            work_book2 = xlwt.Workbook(encoding='utf-8')
            sheet2 = add_sheet(work_book2,'传输距离')
            sheet2.write(0, 0, '传输距离')
            for i in range(len(ll)):
                sheet1.write(0, i + 1, ll[i])
                sheet1.write(i + 1, 0, ll[i])
                sheet2.write(0, i + 1, ll[i])
                sheet2.write(i + 1, 0, ll[i])
            work_book1.save(os.path.join(os.path.abspath('models'), 'CMNET网_metric.xls'))
            work_book2.save(os.path.join(os.path.abspath('models'), 'CMNET网_传输距离.xls'))

        elif pm == 'N1':
            node_dict, hx, lhx, hj = find_node(pm)
            ll = list(node_dict.keys())
            work_book1 = xlwt.Workbook(encoding='utf-8')
            sheet1 = add_sheet(work_book1,'Metric')
            sheet1.write(0, 0, 'Metric')

            work_book2 = xlwt.Workbook(encoding='utf-8')
            sheet2 = add_sheet(work_book2,'传输距离')
            sheet2.write(0, 0, '传输距离')
            for i in range(len(ll)):
                sheet1.write(0, i + 1, ll[i])
                sheet1.write(i + 1, 0, ll[i])
                sheet2.write(0, i + 1, ll[i])
                sheet2.write(i + 1, 0, ll[i])
            work_book1.save(os.path.join(model_path, 'CMNET网_metric.xls'))
            work_book2.save(os.path.join(model_path, 'CMNET网_传输距离.xls'))

        logger.info('上传成功！！！')
        return JsonResponse({'status': 'ok', 'msg': '上传成功！！！'})


# 提供动态表格数据
def show_node(request):
    try:
        if request.method == "GET":

            node_list = CNodes_detail.objects.all().values()
            json_data = list(node_list)

            # print(json_data) # [{'id': 1, 'province': '北京', 'city': '北京', 'office_address1': '三台', 'building_no1': '6B', 'floor1': '2层',...
            # print(type(json_data))
            return JsonResponse(json_data, safe=False)
        else:
            return JsonResponse({'status': 'err', 'msg': '非法请求方式！！'})

    except Exception as e:
        logger.error(e)
        return JsonResponse({'status': 'err', 'msg': e.args})


# 向前端传输Metric矩阵数据
def show_metric(request):
    try:
        if request.method == "GET":
            path = os.path.join(os.path.abspath('excel/cmnetwin'), 'CMNET网工程计算.xls')
            wb = open_exce(path)
            ws = wb.sheet_by_index(0)
            rows = ws.nrows
            cols = ws.ncols
            # 创建一个数组用来存储excel中的数据
            p = []
            for i in range(1, rows):
                d = {}
                for j in range(0, cols):
                    q = '%s' % ws.cell(0, j).value
                    d[q] = ws.cell(i, j).value
                ap = []
                for k, v in d.items():
                    if isinstance(v, float):  # excel中的值默认是float,需要进行判断处理，通过'"%s":%d'，'"%s":"%s"'格式化数组
                        ap.append('"%s":%d' % (k, v))
                    else:
                        ap.append('"%s":"%s"' % (k, v))
                s = '{%s}' % (','.join(ap))  # 继续格式化
                p.append(s)
            t = '[%s]' % (','.join(p))  # 格式化

            # print(json_data)
            # print(type(json_data)) #str
            json_data = json.loads(t)
            # print(json_data, type(json_data)) #list
            # print(t)
            # data = json.dumps(t, ensure_ascii=False)
            # print(data.replace("\\", ""))# "[{"Metric":"北京5","北京5":"","北京6":"","上海5":"","上海6":"","广东5":"","广东6":"","湖北3"。。。

            return JsonResponse(json_data, safe=False)
        else:
            return JsonResponse({'status': 'err', 'msg': '非法请求方式！！'})
    except Exception as e:
        logger.error(e)
        return JsonResponse({'status': 'err', 'msg': e.args})


# 统计多节点
def morenode(request):
    if request.method == 'GET':

        pm = ProjectInfo.objects.filter(createBy=request.user.id, pro_type=2).values('pm')
        # print(pm) # <QuerySet [{'pm': 'N2'}]>

        data = CNodes_detail.objects.filter(plane2=pm, part2__in=['BB', 'BC']).exclude(program='新建').values(
            'province')  # ,city

        prolist = [i['province'] for i in data]

        res = reduce(lambda x, y: x if y in x else x + [y], [[], ] + list(data))  # 格式化

        for i in res:
            i['nums'] = prolist.count(i['province'])  # 节点个数

        json_data = list(filter(lambda x: x['nums'] > 2, res))  # 多节点省份

        # print(json_data) #[{'province': '北京', 'number': 2}, {'province': '上海', 'number': 2},...
        # return JsonResponse(json_data, safe=False)
        return JsonResponse(res, safe=False)

        # return JsonResponse({'status':'ok','msg':'请求成功'})


# 直接上传62x62矩阵方式的点到点流量
def up_flow(request):
    xs = request.POST.get('xs1')
    myFile = request.FILES.get('flowfile')

    if not all([xs, myFile]):
        # 有数据为空
        return JsonResponse({'status': 'err', 'msg': '参数不能为空!'})

    path = settings.MEDIA_ROOT + 'cmnetwin/flow/'  # 上传文件的保存路径，可以自己指定任意的路径
    path2 = os.path.abspath('excel/cmnetwin')

    if not os.path.exists(path):
        logger.info('没有存储文件路径，创建新文件')
        os.makedirs(path)
    if not os.path.exists(path2):
        logger.info('没有存储文件路径，创建新文件')
        os.makedirs(path2)

    with open(os.path.join(path, myFile.name), 'wb+') as f:
        for chunk in myFile.chunks():
            f.write(chunk)

    pro = ProjectInfo()
    pro.id = 2
    pro.createBy_id = request.user.id
    pro.PtoPflowfile = os.path.join(path, myFile.name)
    pro.save()

    excel = xlrd.open_workbook(path + myFile.name)
    book = copy(excel)

    book.save(os.path.join(path2, 'CMNET网_点到点流量.xls'))

    return JsonResponse({'status': 'ok', 'msg': '上传成功！！！'})


'''
# 上传多业务
def up_flows(request):
    xs = request.POST.get('xs1')
    myFile = request.FILES.get('flowfile')

    if not all([xs, myFile]):
        # 有数据为空
        return JsonResponse({'status': 'err', 'msg': '参数不能为空!'})

    path = settings.MEDIA_ROOT + 'cmnetwin/flow/'  # 上传文件的保存路径，可以自己指定任意的路径
    path2 = os.path.abspath('excel/cmnetwin')

    if not os.path.exists(path):
        logger.info('没有存储文件路径，创建新文件')
        os.makedirs(path)
    if not os.path.exists(path2):
        logger.info('没有存储文件路径，创建新文件')
        os.makedirs(path2)

    with open(os.path.join(path, myFile.name), 'wb+') as f:
        for chunk in myFile.chunks():
            f.write(chunk)

    pm = ProjectInfo.objects.get(id=2).pm
    if pm == 'N2':
        sheetname = '流量需求汇总_N2'
    elif pm == 'N1':
        sheetname = '流量需求汇总_N1'

    excel = xlrd.open_workbook(os.path.join(path, myFile.name))
    F_sh = excel.sheet_by_name('F')
    nrow = F_sh.col_values(1)[1:]

    node = []
    for n in nrow:
        if len(n) > 2:
            node.append(n[:2] + '1')
            node.append(n[:2] + '2')
        else:
            node.append(n + '1')
            node.append(n + '2')
    book = copy(excel)
    bs = book.add_sheet(sheetname, cell_overwrite_ok=True)
    rs = excel.sheet_by_index(0).nrows
    cs = excel.sheet_by_index(0).ncols
    sheets = excel.sheet_names()
    # print(sheets)
    fro2 = getSumFiro(sheets)

    for i in range(1, rs):
        bs.write(i, 0, excel.sheet_by_index(0).cell(i, 0).value)
        bs.write(0, i, excel.sheet_by_index(0).cell(i, 0).value)

        for j in range(1, cs):
            bs.write(i, j, xlwt.Formula(getFiro2(fro2, j, i)))

    ns3 = book.add_sheet('点到点流量', cell_overwrite_ok=True)
    for st in range(1, len(node) + 1):
        ns3.write(st, 0, node[st - 1])
        ns3.write(0, st, node[st - 1])
        for en in range(1, len(node) + 1):
            ns3.write(st, en, xlwt.Formula(getFlow(st, en, len(nrow) + 1, sheetname)))

    book.save(os.path.join(path2, 'CMNET网_点到点流量.xls'))

    # .xls转换为xlsx,方便计算利用公式
    # xlstoxlsx(os.path.join(os.getcwd(), 'excel\\cmnetwin\\CMNET网_点到点流量.xls'))

    # filename = os.path.join(os.getcwd(), 'excel\\cmnetwin\\CMNET网_点到点流量.xlsx')
    # 后台自动打开和保存excel文档使文档公式计算生成两套值，方便下面之取值而不取公式
    # just_open(filename)
    # filename2 = os.path.join(os.getcwd(), 'new_book.xlsx')
    # just_open(filename2)
    # wb = load_workbook(filename, data_only=True)
    # wb2 = load_workbook(filename2, data_only=True)
    #
    # ws = wb['点到点流量']
    #
    # ws2 = wb2['点到点流量']
    # # ws1 = wb2.create_sheet("点到点流量备用")
    # # 两个for循环遍历整个excel的单元格内容
    # for i, row in enumerate(ws.iter_rows()):
    #
    #     for j, cell in enumerate(row):
    #         print(cell.value)
    #         ws2.cell(row=i + 1, column=j + 1, value=cell.value)
    #
    # wb2.save(filename2)
    # wb2.close()
    # xlsxtoxls(os.path.join(os.getcwd(), 'new_book.xlsx'))
    return JsonResponse({'status': 'ok', 'msg': '上传成功！！！'})
'''


# 上传多业务
def ntonflow(request):
    if request.method == 'POST':
        xs = request.POST.get('xs')
        myFile = request.FILES.get('myfiles')
        # ratioarr = request.POST.getlist('ratioarr')  # 多节点比例
        # print(ratioarr)

        if not all([xs, myFile]):
            # 有数据为空
            return JsonResponse({'status': 'err', 'msg': '参数不能为空!'})

        path = settings.MEDIA_ROOT + 'cmnetwin/flow/'  # 上传文件的保存路径，可以自己指定任意的路径
        path2 = os.path.abspath('excel/cmnetwin')

        if not os.path.exists(path):
            logger.info('没有存储文件路径，创建新文件')
            os.makedirs(path)
        if not os.path.exists(path2):
            logger.info('没有存储文件路径，创建新文件')
            os.makedirs(path2)

        with open(os.path.join(path, myFile.name), 'wb+') as f:
            for chunk in myFile.chunks():
                f.write(chunk)

        pm = ProjectInfo.objects.get(id=2).pm

        if pm == 'N2':
            sheetname = '流量需求汇总_N2'
        elif pm == 'N1':
            sheetname = '流量需求汇总_N1'
        node_dict, hx, lhx, hj = find_node(pm)
        # print(node_dict)

        excel = xlrd.open_workbook(os.path.join(path, myFile.name))
        # book = copy(excel)
        # sh = book.add_sheet(sheetname, cell_overwrite_ok=True)
        city = excel.sheet_by_index(0).row_values(0)[1:]
        #
        #
        node = list(node_dict.keys())
        # rs = excel.sheet_by_index(0).nrows
        # cs = excel.sheet_by_index(0).ncols
        # sheets = excel.sheet_names()
        #
        # fro2 = getSumFiro(sheets)
        #
        #
        # for i in range(1, rs):
        #     sh.write(i, 0, city[i-1])
        #     sh.write(0, i, city[i-1])
        #
        #     for j in range(1, cs):
        #         sh.write(i, j, xlwt.Formula(getFiro2(fro2, j, i)))
        # wb = open_exce(os.path.join(path2, 'combine.xls'))
        sh = excel.sheet_by_index(0)
        ncols = sh.ncols  # Excel列的数目  原Excel和目标Excel的列表的长度相同
        nrows = sh.nrows
        test = [[0 for c in range(ncols - 1)] for r in range(nrows - 1)]

        for sheetName in excel.sheet_names():
            sheet = excel.sheet_by_name(sheetName)

            for i in range(1, nrows):
                for j in range(1, ncols):

                    ban = sheet.cell(i, j).value

                    if ban == '':
                        ban = 0

                    test[i - 1][j - 1] = test[i - 1][j - 1] + ban
        # print(test)

        book = copy(excel)
        ws = add_sheet(book,sheetname)
        row_0 = sh.row_values(0)  # 获取第一行的值
        col_0 = sh.col_values(0)  # 获取第一列的值
        for i, key in enumerate(row_0):  # 写入新Excel表的第一行
            ws.write(0, i, key)
        for i, key in enumerate(col_0):  # 写入新Excel表的第一列
            ws.write(i, 0, key)
        x = np.array(test)
        # 输出数组的行和列数
        for i in range(x.shape[0]):
            for j in range(x.shape[1]):
                ws.write(i + 1, j + 1, x[i][j])

        p = os.path.join(path2, 'CMNET网_点到点流量.xls')
        p2 = os.path.join(path2, 'new_CMNET网_点到点流量.xls')
        save_exce(book,p,p2)

        # 62x62

        wb = open_exce(os.path.join(path2, 'CMNET网_点到点流量.xls'))
        wb2 = open_exce(os.path.join(path2, 'CMNET网工程计算.xls'))
        new_wb2 = copy(wb2)
        new_ws = add_sheet(new_wb2,'点到点流量')
        ws = wb.sheet_by_name(sheetname)

        if '' in city:
            city.remove('')
        if '合计' in city:
            city.remove('合计')

        data = CNodes_detail.objects.filter(plane2=pm, part2__in=['BB', 'BC']).exclude(program='新建').values('province',
                                                                                                            'city')
        city_dict = {}
        n = []
        m = []
        for i in data:
            city_dict[i['city']] = i['province']

            n.append(i['province'])
            m.append(i['city'])

        T = sorted(set(n),
                   key=n.index)  # 省份['北京', '上海', '广东', '湖北', '辽宁', '江苏', '四川', '陕西', '浙江', '山西', '福建', '山东', '江西', '湖南', '河南', '重庆', '天津', '河北', '内蒙古', '吉林', '黑龙江', '安徽', '广西', '海南', '贵州', '云南', '西藏', '甘肃', '青海', '宁夏', '新疆']
        T1 = sorted(set(m), key=m.index)  # 城市
        c = [i[:2] + str(j + 1) for i in T for j in range(n.count(
            i))]  # ['北京1', '北京2', '上海1', '上海2', '广东1', '广东2', '湖北1', '湖北2', '辽宁1', '辽宁2', '江苏1', '江苏2', '四川1', '四川2', '陕西1'...

        nton = dict(zip(c,
                        node))  # {'北京1': '北京5', '北京2': '北京6', '上海1': '上海5', '上海2': '上海6', '广东1': '广东5', '广东2': '广东6', '湖北1': '湖北3', '湖北2': '湖北4', '辽宁1': '辽宁3', '辽宁2': '辽宁4', '江苏1': '江苏5', '江苏2': '江苏6', '四川1': '四川3', '四川2': '四川4', '陕西1': '陕西3', '陕西2': '陕西4', '浙江1': '浙江5', '浙江2': '浙江6', '山西1': '山西3', '山西2': '山西4', '福建1': '福建3', '福建2': '福建4', '山东1': '山东3', '山东2': '山东4', '江西1': '江西3', '江西2': '江西4', '湖南1': '湖南3', '湖南2': '湖南4', '河南1': '河南3', '河南2': '河南4', '重庆1': '重庆3', '重庆2': '重庆4', '天津1': '天津5', '天津2': '天津6', '河北1': '河北3', '河北2': '河北4', '内蒙1': '内蒙3', '内蒙2': '内蒙4', '吉林1': '吉林3', '吉林2': '吉林4', '黑龙1': '黑龙3', '黑龙2': '黑龙4', '安徽1': '安徽3', '安徽2': '安徽4', '广西1': '广西3', '广西2': '广西4', '海南1': '海南3', '海南2': '海南4', '贵州1': '贵州3', '贵州2': '贵州4', '云南1': '云南3', '云南2': '云南4', '西藏1': '西藏3', '西藏2': '西藏4', '甘肃1': '甘肃3', '甘肃2': '甘肃4', '青海1': '青海3', '青海2': '青海4', '宁夏1': '宁夏3', '宁夏2': '宁夏4', '新疆1': '新疆3', '新疆2': '新疆4'}

        province_all = {}
        for i in range(1, len(city) + 1):
            for j in range(1, len(city) + 1):
                val = ws.cell_value(i, j)
                if i == j:
                    continue
                else:
                    province_all[city[i - 1] + '-' + city[j - 1]] = val
        # print(province_all) #{'北京-上海': 145.37601209068615, '北京-广州': 491.2537317350588, '北京-武汉': 167.1234762744233, '北京-沈阳': 416.17733354238163

        more = {}
        for k, v in province_all.items():
            ll = k.split('-')
            more[city_dict[ll[0]][:2] + str(1) + '-' + city_dict[ll[1]][:2] + str(1)] = v / 2
            more[city_dict[ll[0]][:2] + str(2) + '-' + city_dict[ll[1]][:2] + str(2)] = v / 2
        # print(more) #{'北京1-上海1': 72.68800604534307, '北京2-上海2': 72.68800604534307, '北京1-广东1': 245.626865867
        more_k = list(more.keys())
        for k in more_k:

            for i in range(1, len(node) + 1):
                new_ws.write(i, 0, node[i - 1])
                new_ws.write(0, i, node[i - 1])
                for j in range(1, len(node) + 1):

                    ll = k.split('-')
                    if node[i - 1] == nton[ll[0]] and node[j - 1] == nton[ll[1]]:
                        new_ws.write(i, j, more[k])

        save_exce(new_wb2, os.path.join(path2, 'CMNET网工程计算.xls'), os.path.join(path2, 'new_CMNET网工程计算.xls'))

        return JsonResponse({'status': 'ok', 'msg': '上传成功！！！'})


# 直接导入metric
def up_metric1(request):
    if request.method == 'POST':
        metric_file = request.FILES.get('metric_file')
        if not metric_file:
            return JsonResponse({'status': 'err', 'msg': '不能为空，请上传metric！！！'})
        else:
            try:
                path = settings.MEDIA_ROOT + 'cmnetwin/metric/'  # 上传文件的保存路径，可以自己指定任意的路径
                if not os.path.exists(path):
                    os.makedirs(path)
                with open(os.path.join(path, metric_file.name), 'wb+') as f:
                    for chunk in metric_file.chunks():
                        f.write(chunk)

                #   计算出的临时存放位置

                p = os.path.abspath('excel/cmnetwin')  # 上传文件的保存路径，可以自己指定任意的路径
                if not os.path.exists(p):
                    os.makedirs(p)

                ProjectInfo.objects.filter(createBy=request.user.id, id=2).update(
                    metricfile=os.path.join(path, metric_file.name), pro_type=2)

                excel = xlrd.open_workbook(path + metric_file.name)
                sheet = excel.sheet_by_index(0)  # 获取上传的metric
                pm = ProjectInfo.objects.get(id=2).pm
                node_dict, hx, lhx, hj = find_node(pm)
                nodes = list(node_dict.keys())

                wb = xlwt.Workbook(encoding='utf-8')
                sh = add_sheet(wb,'直联')
                sh.write(0, 0, 'Metric')

                wb2 = xlwt.Workbook(encoding='utf-8')
                sheet1 = add_sheet(wb2,'Metric')
                sheet2 = add_sheet(wb2,'直联')
                sheet1.write(0, 0, 'Metric')

                wb3 = xlwt.Workbook(encoding='utf-8')
                msh = add_sheet(wb3,'Metric')
                msh.write(0, 0, 'Metric')

                nrows = sheet.nrows

                for i in range(1, nrows):
                    sh.write(i, 0, nodes[i - 1])
                    sh.write(0, i, nodes[i - 1])
                    sheet1.write(i, 0, nodes[i - 1])
                    sheet1.write(0, i, nodes[i - 1])
                    sheet2.write(i, 0, nodes[i - 1])
                    sheet2.write(0, i, nodes[i - 1])
                    msh.write(i, 0, nodes[i - 1])
                    msh.write(0, i, nodes[i - 1])
                    for j in range(1, nrows):

                        m = sheet.cell(i, j).value
                        if m == '' or m == 0:
                            m = ''
                        else:

                            sh.write(i, j, m)
                            sh.write(j, i, m)
                            sheet1.write(i, j, m)
                            sheet2.write(i, j, m)
                            sheet2.write(j, i, m)
                            msh.write(i, j, m)

                save_exce(wb,os.path.join(p, 'CMNET网_直联.xls'),os.path.join(p, 'newCMNET网_直联.xls'))
                save_exce(wb2, os.path.join(p, 'CMNET网工程计算.xls'), os.path.join(p, 'newCMNET网工程计算.xls'))
                save_exce(wb3, os.path.join(p, 'CMNET网_metric.xls'), os.path.join(p, 'newCMNET网_metric.xls'))


                str = randomnum()  # 随机数

                metric_num = CMetricTable()
                metric_num.mid = int(str)
                metric_num.fid = int(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                metric_num.file_path = os.path.join(p, 'CMNET网工程计算.xls')
                metric_num.filename = 'CMNET网工程计算'
                metric_num.proid_id = 2
                metric_num.createdBy_id = request.user.id
                metric_num.save()

                return JsonResponse({'status': 'ok', 'msg': '上传成功！！！'})
            except Exception as e:
                logger.info(e)

                return JsonResponse({'status': 'ok', 'msg': '上传失败！！！'})


# metric计算原则一：链路角色+传输距离
def up_metric2(request):
    if request.method == 'POST':
        myFile = request.FILES.get('distance_file')
        js1 = request.POST.get('hx_js')
        js2 = request.POST.get('lhx_hx')
        zs2 = request.POST.get('lhx_hx_zs')
        js3 = request.POST.get('hj_hx')
        zs3 = request.POST.get('hj_hx_zs')
        js4 = request.POST.get('lhx_js')
        zs4 = request.POST.get('lhx_js_zs')
        js5 = request.POST.get('hj_lhx')
        zs5 = request.POST.get('hj_lhx_zs')
        js6 = request.POST.get('hj_js')
        zs6 = request.POST.get('hj_js_zs')
        js7 = request.POST.get('hx_js_2')
        zs7 = request.POST.get('hx_zs2')
        js8 = request.POST.get('lhx_js_2')
        zs8 = request.POST.get('lhx_zs2')
        js9 = request.POST.get('hj_js_2')
        zs9 = request.POST.get('hj_zs2')
        if not all([myFile, js1, js2, zs2, js3, zs3, js4, zs4, js5, zs5, js6, zs6, js7, zs7, js8, zs8, js9, zs9]):
            # 有数据为空
            # print(myFile.name, js1, js2, zs2, js3, zs3, js4, zs4, js5, zs5, js6, zs6, js7, zs7, js8, zs8, js9, zs9)
            return JsonResponse({'status': 'err', 'msg': '参数不能为空!'})

        try:
            path = settings.MEDIA_ROOT + 'cmnetwin/distance/'  # 上传文件的保存路径，可以自己指定任意的路径
            if not os.path.exists(path):
                os.makedirs(path)

            with open(os.path.join(path, myFile.name), 'wb+') as f:
                for chunk in myFile.chunks():
                    f.write(chunk)

            ProjectInfo.objects.filter(createBy=request.user.id, id=2).update(
                distancefile=os.path.join(path, myFile.name))

            excel = xlrd.open_workbook(path + myFile.name)
            sheet = excel.sheet_by_index(0)

            nrows = sheet.nrows

            pm = ProjectInfo.objects.get(id=2).pm

            node_list = CNodes_detail.objects.all()
            if len(node_list) > 0:
                node_dict, hx, lhx, hj = find_node(pm)
                nodes = list(node_dict.keys())
                hx = list(hx.keys())
                lhx = list(lhx.keys())
                hj = list(hj.keys())

                work_book1 = xlwt.Workbook(encoding='utf-8')
                sheet1 = add_sheet(work_book1,'Metric')
                sheet2 = add_sheet(work_book1,'直联')
                wb = xlwt.Workbook(encoding='utf-8')
                sh = add_sheet(wb,'直联')
                sheet1.write(0, 0, 'Metric')

                for i in range(1, nrows):
                    sheet1.write(i, 0, nodes[i - 1])
                    sheet1.write(0, i, nodes[i - 1])
                    sheet2.write(i, 0, nodes[i - 1])
                    sheet2.write(0, i, nodes[i - 1])
                    sh.write(i, 0, nodes[i - 1])
                    sh.write(0, i, nodes[i - 1])
                    for j in range(1, nrows):

                        m = sheet.cell(i, j).value
                        if m == '' or m == 0:
                            m = 0
                        else:
                            if (nodes[i - 1] in hx) and (nodes[j - 1] in hx):
                                if (nodes[i - 1][:2] == nodes[j - 1][:2]):
                                    sheet1.write(i, j, eval(js7) + eval(zs7))
                                    sheet2.write(i, j, eval(js7) + eval(zs7))
                                    sheet2.write(j, i, eval(js7) + eval(zs7))

                                    sh.write(i, j, eval(js7) + eval(zs7))
                                    sh.write(j, i, eval(js7) + eval(zs7))
                                else:
                                    # metric
                                    sheet1.write(i, j, eval(js1))
                                    sheet2.write(i, j, eval(js1))
                                    sheet2.write(j, i, eval(js1))
                                    # 生成计算需要的直联表
                                    sh.write(i, j, eval(js1))
                                    sh.write(j, i, eval(js1))


                            elif ((nodes[i - 1] in hx) and (nodes[j - 1] in lhx)) or (
                                    (nodes[i - 1] in lhx) and (nodes[j - 1] in hx)):

                                sheet1.write(i, j, eval(js2) + m / eval(zs2))
                                sheet2.write(i, j, eval(js2) + m / eval(zs2))
                                sheet2.write(j, i, eval(js2) + m / eval(zs2))

                                sh.write(i, j, eval(js2) + m / eval(zs2))
                                sh.write(j, i, eval(js2) + m / eval(zs2))

                            elif ((nodes[i - 1] in hj) and (nodes[j - 1] in hx)) or (
                                    (nodes[i - 1] in hx) and (nodes[j - 1] in hj)):

                                sheet1.write(i, j, eval(js3) + m / eval(zs3))
                                sheet2.write(i, j, eval(js3) + m / eval(zs3))
                                sheet2.write(j, i, eval(js3) + m / eval(zs3))

                                sh.write(i, j, eval(js3) + m / eval(zs3))
                                sh.write(j, i, eval(js3) + m / eval(zs3))

                            elif (nodes[i - 1] in lhx) and (nodes[j - 1] in lhx):
                                if (nodes[i - 1][:2] == nodes[j - 1][:2]):
                                    sheet1.write(i, j, eval(js8) + eval(zs8))
                                    sheet2.write(i, j, eval(js8) + eval(zs8))
                                    sheet2.write(j, i, eval(js8) + eval(zs8))

                                    sh.write(i, j, eval(js8) + eval(zs8))
                                    sh.write(j, i, eval(js8) + eval(zs8))
                                else:
                                    # metric
                                    sheet1.write(i, j, eval(js4) + m / eval(zs4))
                                    sheet2.write(i, j, eval(js4) + m / eval(zs4))
                                    sheet2.write(j, i, eval(js4) + m / eval(zs4))
                                    # 生成计算需要的直联表
                                    sh.write(i, j, eval(js4) + m / eval(zs4))
                                    sh.write(j, i, eval(js4) + m / eval(zs4))

                            elif ((nodes[i - 1] in hj) and (nodes[j - 1] in lhx)) or (
                                    (nodes[i - 1] in lhx) and (nodes[j - 1] in hj)):

                                sheet1.write(i, j, eval(js5) + m / eval(zs5))
                                sheet2.write(i, j, eval(js5) + m / eval(zs5))
                                sheet2.write(j, i, eval(js5) + m / eval(zs5))

                                sh.write(i, j, eval(js5) + m / eval(zs5))
                                sh.write(j, i, eval(js5) + m / eval(zs5))

                            elif (nodes[i - 1] in hj) and (nodes[j - 1] in hj):
                                if (nodes[i - 1][:2] == nodes[j - 1][:2]):
                                    sheet1.write(i, j, eval(js9) + eval(zs9))
                                    sheet2.write(i, j, eval(js9) + eval(zs9))
                                    sheet2.write(j, i, eval(js9) + eval(zs9))

                                    sh.write(i, j, eval(js9) + eval(zs9))
                                    sh.write(j, i, eval(js9) + eval(zs9))
                                else:
                                    # metric
                                    sheet1.write(i, j, eval(js6) + m / eval(zs6))
                                    sheet2.write(i, j, eval(js6) + m / eval(zs6))
                                    sheet2.write(j, i, eval(js6) + m / eval(zs6))
                                    # 生成计算需要的直联表
                                    sh.write(i, j, eval(js6) + m / eval(zs6))
                                    sh.write(j, i, eval(js6) + m / eval(zs6))

                p1 = os.path.abspath('excel/cmnetwin/metric')  # 上传文件的保存路径，可以自己指定任意的路径
                if not os.path.exists(p1):
                    os.makedirs(path)

                save_exce(work_book1,os.path.join(p1, 'CMNET网_metric.xls'),os.path.join(p1, 'newCMNET网_metric.xls'))


                # 保存计算直联
                path1 = os.path.abspath('excel/cmnetwin')
                if not os.path.exists(path1):
                    os.makedirs(path1)

                save_exce(wb,os.path.join(path1, 'CMNET网_直联.xls'),os.path.join(path1, 'CMNET网_直联2.xls'))


                str = randomnum()  # 随机数

                metric_num = CMetricTable()
                metric_num.mid = int(str)
                metric_num.fid = int(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                metric_num.file_path = os.path.join(p1, 'CMNET网工程计算.xls')
                metric_num.filename = 'CMNET网工程计算'
                metric_num.proid_id = 2
                metric_num.createdBy_id = request.user.id
                metric_num.save()

                return JsonResponse({'status': 'ok', 'msg': '上传成功！'})
        except Exception as e:
            logger.info(e)
            return JsonResponse({'status': 'ok', 'msg': '上传失败！'})


# metric计算原则二：传输距离
def up_metric3(request):
    if request.method == 'POST':
        myFile = request.FILES.get('distance_file2')
        zs1 = request.POST.get('hx_js2_zs')
        zs2 = request.POST.get('lhx_hx2_zs')
        zs3 = request.POST.get('hj_hx2_zs')
        zs4 = request.POST.get('lhx_js2_zs')
        zs5 = request.POST.get('hj_lhx2_zs')
        zs6 = request.POST.get('hj_js2_zs')
        js7 = request.POST.get('hx_js2_2')
        zs7 = request.POST.get('hx_js2_zs2')
        js8 = request.POST.get('lhx_js2_2')
        zs8 = request.POST.get('lhx_js2_zs2')
        js9 = request.POST.get('hj_js2_2')
        zs9 = request.POST.get('hj_js2_zs2')

        try:
            pm = ProjectInfo.objects.get(id=2).pm

            if not all([myFile, zs1, zs2, zs3, zs4, zs5, zs6, js7, zs7, js8, zs8, js9, zs9]):
                # print(myFile.name, zs1, zs2, zs3, zs4, zs5, zs6, js7, zs7, js8, zs8, js9, zs9)
                return JsonResponse({'status': 'err', 'msg': '参数不能为空！'})
            p1 = settings.MEDIA_ROOT + 'cmnetwin/distance/'  # 上传文件的保存路径，可以自己指定任意的路径
            if not os.path.exists(p1):
                os.makedirs(p1)

            with open(os.path.join(p1, myFile.name), 'wb+') as f:
                for chunk in myFile.chunks():
                    f.write(chunk)

            ProjectInfo.objects.filter(createBy=request.user.id, id=2).update(
                distancefile=os.path.join(p1, myFile.name))

            excel = xlrd.open_workbook(p1 + myFile.name)
            sheet = excel.sheet_by_index(0)

            nrows = sheet.nrows

            node_list = CNodes_detail.objects.all()
            if len(node_list) > 0:
                node_dict, hx, lhx, hj = find_node(pm)
                nodes = list(node_dict.keys())
                hx = list(hx.keys())
                lhx = list(lhx.keys())
                hj = list(hj.keys())

                work_book1 = xlwt.Workbook(encoding='utf-8')
                sheet1 = work_book1.add_sheet('Metric', cell_overwrite_ok=True)
                sheet2 = work_book1.add_sheet('直联', cell_overwrite_ok=True)
                sheet1.write(0, 0, 'Metric')
                wb = xlwt.Workbook(encoding='utf-8')
                sh = wb.add_sheet('直联', cell_overwrite_ok=True)

                for i in range(1, nrows):
                    sheet1.write(i, 0, nodes[i - 1])
                    sheet1.write(0, i, nodes[i - 1])
                    sheet2.write(i, 0, nodes[i - 1])
                    sheet2.write(0, i, nodes[i - 1])
                    sh.write(i, 0, nodes[i - 1])
                    sh.write(0, i, nodes[i - 1])
                    for j in range(1, nrows):

                        m = sheet.cell(i, j).value
                        if m == '' or m == 0:
                            pass
                        else:

                            if (nodes[i - 1] in hx) and (nodes[j - 1] in hx):
                                if (nodes[i - 1][:2] == nodes[j - 1][:2]):
                                    sheet1.write(i, j, eval(js7) + eval(zs7))
                                    sheet2.write(i, j, eval(js7) + eval(zs7))
                                    sheet2.write(j, i, eval(js7) + eval(zs7))

                                    sh.write(i, j, eval(js7) + eval(zs7))
                                    sh.write(j, i, eval(js7) + eval(zs7))
                                else:
                                    # metric
                                    sheet1.write(i, j, m / eval(zs1))
                                    sheet2.write(i, j, m / eval(zs1))
                                    sheet2.write(j, i, m / eval(zs1))
                                    # 生成计算需要的直联表
                                    sh.write(i, j, m / eval(zs1))
                                    sh.write(j, i, m / eval(zs1))


                            elif ((nodes[i - 1] in hx) and (nodes[j - 1] in lhx)) or (
                                    (nodes[i - 1] in lhx) and (nodes[j - 1] in hx)):

                                sheet1.write(i, j, m / eval(zs2))
                                sheet2.write(i, j, m / eval(zs2))
                                sheet2.write(j, i, m / eval(zs2))

                                sh.write(i, j, m / eval(zs2))
                                sh.write(j, i, m / eval(zs2))

                            elif ((nodes[i - 1] in hj) and (nodes[j - 1] in hx)) or (
                                    (nodes[i - 1] in hx) and (nodes[j - 1] in hj)):

                                sheet1.write(i, j, m / eval(zs3))
                                sheet2.write(i, j, m / eval(zs3))
                                sheet2.write(j, i, m / eval(zs3))

                                sh.write(i, j, m / eval(zs3))
                                sh.write(j, i, m / eval(zs3))

                            elif (nodes[i - 1] in lhx) and (nodes[j - 1] in lhx):
                                if (nodes[i - 1][:2] == nodes[j - 1][:2]):
                                    sheet1.write(i, j, eval(js8) + eval(zs8))
                                    sheet2.write(i, j, eval(js8) + eval(zs8))
                                    sheet2.write(j, i, eval(js8) + eval(zs8))

                                    sh.write(i, j, eval(js8) + eval(zs8))
                                    sh.write(j, i, eval(js8) + eval(zs8))
                                else:
                                    # metric
                                    sheet1.write(i, j, m / eval(zs4))
                                    sheet2.write(i, j, m / eval(zs4))
                                    sheet2.write(j, i, m / eval(zs4))
                                    # 生成计算需要的直联表
                                    sh.write(i, j, m / eval(zs4))
                                    sh.write(j, i, m / eval(zs4))

                            elif ((nodes[i - 1] in hj) and (nodes[j - 1] in lhx)) or (
                                    (nodes[i - 1] in lhx) and (nodes[j - 1] in hj)):

                                sheet1.write(i, j, m / eval(zs5))
                                sheet2.write(i, j, m / eval(zs5))
                                sheet2.write(j, i, m / eval(zs5))

                                sh.write(i, j, m / eval(zs5))
                                sh.write(j, i, m / eval(zs5))

                            elif (nodes[i - 1] in hj) and (nodes[j - 1] in hj):
                                if (nodes[i - 1][:2] == nodes[j - 1][:2]):
                                    sheet1.write(i, j, eval(js9) + eval(zs9))
                                    sheet2.write(i, j, eval(js9) + eval(zs9))
                                    sheet2.write(j, i, eval(js9) + eval(zs9))

                                    sh.write(i, j, eval(js9) + eval(zs9))
                                    sh.write(j, i, eval(js9) + eval(zs9))
                                else:
                                    # metric
                                    sheet1.write(i, j, m / eval(zs6))
                                    sheet2.write(i, j, m / eval(zs6))
                                    sheet2.write(j, i, m / eval(zs6))
                                    # 生成计算需要的直联表
                                    sh.write(i, j, m / eval(zs6))
                                    sh.write(j, i, m / eval(zs6))

                path = os.path.abspath('excel/cmnetwin/metric')  # 上传文件的保存路径，可以自己指定任意的路径
                if not os.path.exists(path):
                    os.makedirs(path)

                if os.path.exists(os.path.join(path, 'CMNET网_metric.xls')):
                    os.remove(os.path.join(path, 'CMNET网_metric.xls'))
                    work_book1.save(os.path.join(path, 'CMNET网_metric.xls'))
                else:
                    work_book1.save(os.path.join(path, 'CMNET网_metric.xls'))

                # 保存计算直联
                path1 = os.path.abspath('excel/cmnetwin')
                if not os.path.exists(path1):
                    os.makedirs(path1)

                if os.path.exists(os.path.join(path1, 'CMNET网_直联.xls')):
                    wb.save(
                        os.path.join(path1, 'CMNET网_直联2.xls'))  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
                    os.remove(os.path.join(path1, 'CMNET网_直联.xls'))
                    os.rename(os.path.join(path1, 'CMNET网_直联2.xls'), os.path.join(path1, 'CMNET网_直联.xls'))
                else:
                    wb.save(os.path.join(path1, 'CMNET网_直联.xls'))

                str = randomnum()  # 随机数

                metric_num = CMetricTable()
                metric_num.mid = int(str)
                metric_num.fid = int(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                metric_num.file_path = os.path.join(path, 'CMNET网工程计算.xls')
                metric_num.filename = 'CMNET网工程计算'
                metric_num.proid_id = 2
                metric_num.createdBy_id = request.user.id
                metric_num.save()

                return JsonResponse({'status': 'ok', 'msg': '上传成功！'})
        except Exception as e:
            logger.info(e)
            return JsonResponse({'status': 'ok', 'msg': '上传失败！'})


# metric计算原则三
def up_metric4(request):
    if request.method == 'POST':
        myFile = request.FILES.get('direction_file')
        js1 = request.POST.get('hx_js3')
        js2 = request.POST.get('lhx_hx3')
        js3 = request.POST.get('hj_hx3')
        js4 = request.POST.get('lhx_js3')
        js5 = request.POST.get('hj_lhx3')
        js6 = request.POST.get('hj_js3')
        js7 = request.POST.get('hx_js3_2')
        zs7 = request.POST.get('hx_js3_zs2')
        js8 = request.POST.get('lhx_js3_2')
        zs8 = request.POST.get('lhx_js3_zs2')
        js9 = request.POST.get('hj_js3_2')
        zs9 = request.POST.get('hj_js3_zs2')

        try:

            pm = ProjectInfo.objects.get(id=2).pm

            if not all([myFile, js1, js2, js3, js4, js5, js6, js7, zs7, js8, zs8, js9, zs9]):
                return JsonResponse({'status': 'err', 'msg': '参数不能为空！'})
            p = settings.MEDIA_ROOT + 'cmnetwin/direction/'  # 上传中继方向文件的保存路径，可以自己指定任意的路径
            if not os.path.exists(p):
                os.makedirs(p)

            ProjectInfo.objects.filter(createBy=request.user.id, i=2).update(
                Relaydirectionfile=os.path.join(p, myFile.name))

            with open(os.path.join(p, myFile.name), 'wb+') as f:
                for chunk in myFile.chunks():
                    f.write(chunk)

            excel = open_exce(p + myFile.name)
            sheet = excel.sheet_by_index(0)

            nrows = sheet.nrows

            node_list = CNodes_detail.objects.all()
            if len(node_list) > 0:
                node_dict, hx, lhx, hj = find_node(pm)
                nodes = list(node_dict.keys())
                hx = list(hx.keys())
                lhx = list(lhx.keys())
                hj = list(hj.keys())

                work_book1 = xlwt.Workbook(encoding='utf-8')
                sheet1 = work_book1.add_sheet('Metric', cell_overwrite_ok=True)
                sheet2 = work_book1.add_sheet('直联', cell_overwrite_ok=True)
                sheet1.write(0, 0, 'Metric')
                wb = xlwt.Workbook(encoding='utf-8')
                sh = wb.add_sheet('直联', cell_overwrite_ok=True)

                for i in range(1, nrows):
                    sheet1.write(i, 0, nodes[i - 1])
                    sheet1.write(0, i, nodes[i - 1])
                    sheet2.write(i, 0, nodes[i - 1])
                    sheet2.write(0, i, nodes[i - 1])
                    sh.write(i, 0, nodes[i - 1])
                    sh.write(0, i, nodes[i - 1])
                    for j in range(1, nrows):

                        m = sheet.cell(i, j).value
                        if m == '' or m == 0:
                            pass
                        else:
                            if (nodes[i - 1] in hx) and (nodes[j - 1] in hx):
                                if (nodes[i - 1][:2] == nodes[j - 1][:2]):
                                    sheet1.write(i, j, eval(js7) + eval(zs7))
                                    sheet2.write(i, j, eval(js7) + eval(zs7))
                                    sheet2.write(j, i, eval(js7) + eval(zs7))

                                    sh.write(i, j, eval(js7) + eval(zs7))
                                    sh.write(j, i, eval(js7) + eval(zs7))
                                else:
                                    # metric
                                    sheet1.write(i, j, eval(js1))
                                    sheet2.write(i, j, eval(js1))
                                    sheet2.write(j, i, eval(js1))
                                    # 生成计算需要的直联表
                                    sh.write(i, j, eval(js1))
                                    sh.write(j, i, eval(js1))


                            elif ((nodes[i - 1] in hx) and (nodes[j - 1] in lhx)) or (
                                    (nodes[i - 1] in lhx) and (nodes[j - 1] in hx)):

                                sheet1.write(i, j, eval(js2))
                                sheet2.write(i, j, eval(js2))
                                sheet2.write(j, i, eval(js2))

                                sh.write(i, j, eval(js2))
                                sh.write(j, i, eval(js2))

                            elif ((nodes[i - 1] in hj) and (nodes[j - 1] in hx)) or (
                                    (nodes[i - 1] in hx) and (nodes[j - 1] in hj)):

                                sheet1.write(i, j, eval(js3))
                                sheet2.write(i, j, eval(js3))
                                sheet2.write(j, i, eval(js3))

                                sh.write(i, j, eval(js3))
                                sh.write(j, i, eval(js3))

                            elif (nodes[i - 1] in lhx) and (nodes[j - 1] in lhx):
                                if (nodes[i - 1][:2] == nodes[j - 1][:2]):
                                    sheet1.write(i, j, eval(js8) + eval(zs8))
                                    sheet2.write(i, j, eval(js8) + eval(zs8))
                                    sheet2.write(j, i, eval(js8) + eval(zs8))

                                    sh.write(i, j, eval(js8) + eval(zs8))
                                    sh.write(j, i, eval(js8) + eval(zs8))
                                else:

                                    # metric
                                    sheet1.write(i, j, eval(js4))
                                    sheet2.write(i, j, eval(js4))
                                    sheet2.write(j, i, eval(js4))
                                    # 生成计算需要的直联表
                                    sh.write(i, j, eval(js4))
                                    sh.write(j, i, eval(js4))

                            elif ((nodes[i - 1] in hj) and (nodes[j - 1] in lhx)) or (
                                    (nodes[i - 1] in lhx) and (nodes[j - 1] in hj)):

                                sheet1.write(i, j, eval(js5))
                                sheet2.write(i, j, eval(js5))
                                sheet2.write(j, i, eval(js5))

                                sh.write(i, j, eval(js5))
                                sh.write(j, i, eval(js5))

                            elif (nodes[i - 1] in hj) and (nodes[j - 1] in hj):

                                if (nodes[i - 1][:2] == nodes[j - 1][:2]):
                                    sheet1.write(i, j, eval(js9) + eval(zs9))
                                    sheet2.write(i, j, eval(js9) + eval(zs9))
                                    sheet2.write(j, i, eval(js9) + eval(zs9))

                                    sh.write(i, j, eval(js9) + eval(zs9))
                                    sh.write(j, i, eval(js9) + eval(zs9))
                                else:

                                    # metric
                                    sheet1.write(i, j, eval(js6))
                                    sheet2.write(i, j, eval(js6))
                                    sheet2.write(j, i, eval(js6))
                                    # 生成计算需要的直联表
                                    sh.write(i, j, eval(js6))
                                    sh.write(j, i, eval(js6))

                path = os.path.abspath('excel/cmnetwin/metric')  # 上传文件的保存路径，可以自己指定任意的路径
                if not os.path.exists(path):
                    os.makedirs(path)

                if os.path.exists(os.path.join(path, 'CMNET网_metric.xls')):
                    os.remove(os.path.join(path, 'CMNET网_metric.xls'))
                    work_book1.save(os.path.join(path, 'CMNET网_metric.xls'))
                else:
                    work_book1.save(os.path.join(path, 'CMNET网_metric.xls'))

                # 保存计算直联
                path1 = os.path.abspath('excel/cmnetwin')
                if not os.path.exists(path1):
                    os.makedirs(path1)

                if os.path.exists(os.path.join(path1, 'CMNET网_直联.xls')):
                    wb.save(
                        os.path.join(path1, 'CMNET网_直联2.xls'))  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
                    os.remove(os.path.join(path1, 'CMNET网_直联.xls'))
                    os.rename(os.path.join(path1, 'CMNET网_直联2.xls'), os.path.join(path1, 'CMNET网_直联.xls'))
                else:
                    wb.save(os.path.join(path1, 'CMNET网_直联.xls'))

                str = randomnum()  # 随机数

                metric_num = CMetricTable()
                metric_num.mid = int(str)
                metric_num.fid = int(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                metric_num.file_path = os.path.join(path, 'CMNET网工程计算.xls')
                metric_num.filename = 'CMNET网工程计算'
                metric_num.proid_id = 2
                metric_num.createdBy_id = request.user.id
                metric_num.save()

            return JsonResponse({'status': 'ok', 'msg': '上传成功！！！'})
        except Exception as e:
            logger.info(e)
            return JsonResponse({'status': 'ok', 'msg': '上传成功！！！'})


# 导出metric
def down_metric1(request):
    if request.method == 'GET':
        path = os.path.abspath('excel/cmnetwin/metric')
        new_path = os.path.join(path, 'CMNET网_metric.xls')

        if not new_path:
            return JsonResponse({'status': 'err', 'msg': '没有计算得到的metric值结果表'})

        wb = open_exce(new_path)
        wb2 = copy(wb)

        response = HttpResponse(content_type='application/octet-stream')
        response['Content-Disposition'] = 'attachment; filename="{0}"'.format(escape_uri_path('CMNET网_metric.xls'))
        wb2.save(response)
        return response


# 获取节点
def get_nodes(pro_type):
    if pro_type == '2':
        p1 = os.path.join(os.path.abspath('excel/cmnetwin'), 'CMNET网_直联.xls')
        if os.path.exists(p1):

            book = open_exce(p1)
            sheet = book.sheet_by_name('直联')
            nodes = sheet.row_values(0)[1:]

            remove_list = ['源节点', '目的节点', '合计', '节点发送带宽和']
            for i in remove_list:
                if i in nodes:
                    nodes.remove(i)
            num = len(nodes)
            return nodes, num
    else:

        p2 = os.path.join(os.path.abspath('excel/ipwin'), 'IP网_直联.xls')
        if os.path.exists(p2):
            book = open_exce(p2)
            sheet = book.sheet_by_name('直联')
            nodes = sheet.row_values(0)[1:]

            remove_list = ['源节点', '目的节点', '合计', '节点发送带宽和']
            for i in remove_list:
                if i in nodes:
                    nodes.remove(i)
            num = len(nodes)
            return nodes, num


# 稳态初始化点点路由
def initRouteExcel(pro_type):
    if pro_type == '2':
        # 初始化存储位置及文件名
        p = os.path.join(os.path.abspath('excel/cmnetwin'), 'CMNET网工程计算.xls')
        p2 = os.path.join(os.path.abspath('excel/cmnetwin'), 'newCMNET网工程计算.xls')

        book = open_exce(p)  # 打开文件，并且保留原格式
        book2 = copy(book)

        sheet = book.sheet_by_name('直联')  # 打开sheet页
        pm = ProjectInfo.objects.get(id=2).pm

        node_dict, hx, lhx, hj = find_node(pm)

        nodes = list(node_dict.keys())
        num = len(nodes)

        sheet1 = book2.add_sheet('点点路由', cell_overwrite_ok=True)

        for j in range(1, num + 1):
            start_node = sheet.cell_value(j, 0)

            for z in range(1, num + 1):
                end_node = sheet.cell_value(0, z)
                sheet1.write((j - 1) * num + z, 0, start_node + '-' + end_node)
                if j == z:
                    sheet1.write((j - 1) * num + z, 1, 'NULL')
                    sheet1.write((j - 1) * num + z, 2, 'NULL')
                    sheet1.write((j - 1) * num + z, 3, 'NULL')
                    sheet1.write((j - 1) * num + z, 4, 'NULL')

        sheet1.write(0, 0, '源节点-目的节点')
        sheet1.write(0, 1, '同权路由数量')
        sheet1.write(0, 2, '度量值')
        sheet1.write(0, 3, '路由跳数')
        sheet1.write(0, 4, '路由')

        if os.path.exists(p):
            book2.save(p2)  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
            os.remove(p)
            os.rename(p2, p)
        else:
            book2.save(p)  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的

    elif pro_type == '1':
        book = open_exce(os.path.join(os.path.abspath('excel/ipwin'), 'IP网_直联.xls'))  # 打开文件
        sheet = book.sheet_by_index(0)  # 打开sheet页
        book2 = xlwt.Workbook(encoding='utf-8')
        sh1 = book2.add_sheet('直联', cell_overwrite_ok=True)
        nodes, num = get_nodes(pro_type)

        for c in range(1, num + 1):
            node = sheet.cell(c, 0).value
            sh1.write(c, 0, node)
            sh1.write(0, c, node)

            for r in range(1, num + 1):
                v = sheet.cell(c, r).value
                sh1.write(c, r, v)

        sheet1 = book2.add_sheet('点点路由', cell_overwrite_ok=True)
        for j in range(1, num + 1):
            start_node = sheet.cell_value(j, 0)

            for z in range(1, num + 1):
                end_node = sheet.cell_value(0, z)
                sheet1.write((j - 1) * num + z, 0, start_node + '-' + end_node)
                if j == z:
                    sheet1.write((j - 1) * num + z, 1, 'NULL')
                    sheet1.write((j - 1) * num + z, 2, 'NULL')
                    sheet1.write((j - 1) * num + z, 3, 'NULL')
                    sheet1.write((j - 1) * num + z, 4, 'NULL')

        sheet1.write(0, 0, '源节点-目的节点')
        sheet1.write(0, 1, '同权路由数量')
        sheet1.write(0, 2, '度量值')
        sheet1.write(0, 3, '路由跳数')
        sheet1.write(0, 4, '路由')

        p = os.path.join(os.path.abspath('excel/ipwin'), 'new_book.xls')
        p2 = os.path.join(os.path.abspath('excel/ipwin'), 'book_new.xls')
        if os.path.exists(p):
            os.remove(p)
            book2.save(p2)  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
            os.rename(p2, p)
        else:
            book2.save(p)  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
    else:
        logger.info('输入错误！！！')


# 稳态调用计算点点路由
def import_excel(request):
    if request.method == 'GET':
        pro_type = request.GET.get('pro_type')

        if pro_type == '2':
            path1 = os.path.join(os.path.abspath('excel/cmnetwin'), 'CMNET网工程计算.xls')
            path2 = os.path.join(os.path.abspath('excel/cmnetwin'), 'newCMNET网工程计算.xls')

            initRouteExcel(pro_type)
            book = open_exce(path1)

            sheet_list = book.sheet_names()
            sheet = book.sheet_by_name('直联')
            sheet1 = book.sheet_by_name('点点路由')
            book2 = copy(book)

            sheet_index = sheet_list.index('点点路由')
            sh = book2.get_sheet(sheet_index)

            pm = ProjectInfo.objects.get(id=2).pm

            node_dict, hx, lhx, hj = find_node(pm)

            nodes = list(node_dict.keys())
            num = len(nodes)

            nlist = []
            for i in range(1, num + 1):
                elist = []
                for j in range(1, num + 1):

                    bandwidth = sheet.cell_value(i, j)

                    if i == j or bandwidth == '' or bandwidth == 0:
                        continue
                    else:
                        edge = Edge(nodes[i - 1], nodes[j - 1], bandwidth)

                        if nodes[i - 1] == edge.startNodeId:
                            elist.append(edge)

                        node = Node(nodes[i - 1], elist)

                    nlist.append(node)

            graph = Graph(nlist)

            for startNodeId in nodes:

                for endNodeId in nodes:

                    if startNodeId != endNodeId:

                        originNodeId, destNodeId, route, weight = graph.dijkstra(startNodeId, endNodeId)

                        for j in range(1, num + 1):
                            for z in range(1, num + 1):
                                StoE = sheet1.cell((j - 1) * num + z, 0).value
                                if originNodeId == destNodeId:
                                    sh.write((j - 1) * num + z, 1, 'NULL')
                                    sh.write((j - 1) * num + z, 2, 'NULL')
                                    sh.write((j - 1) * num + z, 3, 'NULL')
                                    sh.write((j - 1) * num + z, 4, 'NULL')
                                if originNodeId + "-" + destNodeId == StoE:
                                    sh.write((j - 1) * num + z, 1, len(route))
                                    sh.write((j - 1) * num + z, 2, weight)

                                    for p in route:
                                        pa = p + [destNodeId]
                                        sh.write((j - 1) * num + z, 3, len(pa) - 1)
                                        sh.write((j - 1) * num + z, 4 + route.index(p), '-'.join(pa))

                        print(originNodeId, destNodeId, '计算中。。。。。')

            if os.path.exists(path1):

                book2.save(path2)
                os.remove(path1)
                os.rename(path2, path1)
            else:
                book2.save(path1)
            bandwidth_allocation(pro_type, path1, path2)
            return JsonResponse({'status': 'ok', 'msg': '计算完成'})
        elif pro_type == '1':
            path1 = os.path.join(os.path.abspath('excel/ipwin'), 'new_book.xls')
            path2 = os.path.join(os.path.abspath('excel/ipwin'), 'book_new.xls')
            if not os.path.exists(path1):
                initRouteExcel(pro_type)
                book = open_exce(path1)
            else:

                book = open_exce(path1)

            sheet_list = book.sheet_names()
            sheet = book.sheet_by_name('直联')
            sheet1 = book.sheet_by_name('点点路由')
            book2 = copy(book)

            sheet_index = sheet_list.index('点点路由')
            sh = book2.get_sheet(sheet_index)

            nodes, num = get_nodes(pro_type)

            nlist = []
            for i in range(1, num + 1):
                elist = []
                for j in range(1, num + 1):

                    bandwidth = sheet.cell_value(i, j)

                    if i == j or bandwidth == '' or bandwidth == 0:
                        continue
                    else:
                        edge = Edge(nodes[i - 1], nodes[j - 1], bandwidth)

                        if nodes[i - 1] == edge.startNodeId:
                            elist.append(edge)

                        node = Node(nodes[i - 1], elist)

                    nlist.append(node)

            graph = Graph(nlist)
            for startNodeId in nodes:

                for endNodeId in nodes:

                    if startNodeId != endNodeId:
                        try:

                            originNodeId, destNodeId, route, weight = graph.dijkstra(startNodeId, endNodeId)
                            for j in range(1, num + 1):
                                for z in range(1, num + 1):
                                    StoE = sheet1.cell((j - 1) * num + z, 0).value
                                    if originNodeId == destNodeId:
                                        sh.write((j - 1) * num + z, 1, 'NULL')
                                        sh.write((j - 1) * num + z, 2, 'NULL')
                                        sh.write((j - 1) * num + z, 3, 'NULL')
                                        sh.write((j - 1) * num + z, 4, 'NULL')

                                    if originNodeId + "-" + destNodeId == StoE:
                                        sh.write((j - 1) * num + z, 1, len(route))
                                        sh.write((j - 1) * num + z, 2, weight)

                                        for p in route:
                                            pa = p + [destNodeId]
                                            sh.write((j - 1) * num + z, 3, len(pa) - 1)
                                            sh.write((j - 1) * num + z, 4 + route.index(p), '-'.join(pa))

                            print(originNodeId, destNodeId, '计算中。。。。。')
                        except Exception as e:
                            logger.error(e)

            if os.path.exists(path1):
                try:
                    os.remove(path1)
                    book2.save(path2)

                    os.rename(path2, path1)
                except PermissionError:
                    return JsonResponse({'status': 'fail', 'msg': '文件被打开，请关闭文件，重新计算！！！'})
            else:
                book2.save(path1)

            bandwidth_allocation(pro_type, path1, path2)

            return JsonResponse({'status': 'ok', 'msg': '计算完成'})
    else:

        return JsonResponse({'status': 'fail', 'msg': '错误请求，计算失败'})


# 节点故障（核心层，类核心，全网）
def gz(request):
    if request.method == 'POST':
        check_box_list = request.POST.getlist('ids')
        pro_type = request.POST.get('pro_type')

        path = os.path.abspath('excel/cmnetwin/max/')
        if not os.path.exists(path):
            os.mkdir(path)

        pm = ProjectInfo.objects.get(id=2).pm

        node_dict, hx, lhx, hj = find_node(pm)

        if check_box_list == ['1,2']:

            hx = list(hx.keys())
            lhx = list(lhx.keys())
            hx_node = hx + lhx
            lhx_fn(hx_node, pro_type)

            all_exce = glob.glob(path + "/*.xls")
            print("该目录下有" + str(len(all_exce)) + "个excel文件：")

            for exce in all_exce:
                fh = open_exce(exce)
                sheet = fh.sheet_by_name('直联')  # 打开sheet页
                sheet1 = fh.sheet_by_name('点点路由')
                sheet_list = fh.sheet_names()
                nodes = list(node_dict.keys())
                num = len(nodes)

                book2 = copy(fh)

                sheet_index = sheet_list.index('点点路由')
                sh = book2.get_sheet(sheet_index)

                logger.info(exce, '进行最短寻路中。。。')
                nlist = []

                for i in range(1, num + 1):
                    elist = []
                    for j in range(1, num + 1):

                        bandwidth = sheet.cell_value(i, j)

                        if i == j or bandwidth == '' or bandwidth == 0:
                            continue
                        else:
                            edge = Edge(nodes[i - 1], nodes[j - 1], bandwidth)

                            if nodes[i - 1] == edge.startNodeId:
                                elist.append(edge)
                    if elist == []:
                        continue
                    else:
                        node = Node(nodes[i - 1], elist)
                    nlist.append(node)
                graph = Graph(nlist)
                name = os.listdir(path)

                for na in name:
                    if na[:-4] == exce[-7:-4]:
                        na = exce[-7:-4]

                        for i in nodes:

                            startNodeId = i

                            for j in nodes:
                                endNodeId = j
                                if startNodeId == na or endNodeId == na:
                                    continue

                                else:

                                    originNodeId, destNodeId, route, weight = graph.dijkstra(startNodeId, endNodeId)

                                for j in range(1, num + 1):
                                    for z in range(1, num + 1):
                                        StoE = sheet1.cell((j - 1) * num + z, 0).value
                                        if originNodeId == destNodeId:
                                            sh.write((j - 1) * num + z, 1, 'NULL')
                                            sh.write((j - 1) * num + z, 2, 'NULL')
                                            sh.write((j - 1) * num + z, 3, 'NULL')
                                            sh.write((j - 1) * num + z, 4, 'NULL')

                                        if StoE == originNodeId + "-" + destNodeId:
                                            sh.write((j - 1) * num + z, 1, len(route))
                                            sh.write((j - 1) * num + z, 2, weight)

                                            for p in route:
                                                pa = p + [destNodeId]
                                                sh.write((j - 1) * num + z, 3, len(pa) - 1)
                                                sh.write((j - 1) * num + z, 4 + route.index(p), '-'.join(pa))
                                logger.info(originNodeId, destNodeId, '计算中。。。。。')
                if os.path.exists(exce):
                    book2.save(os.path.join(path, 'book_new.xls'))  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
                    os.remove(exce)
                    os.rename(os.path.join(path, 'book_new.xls'), exce)
                else:
                    book2.save(exce)
                bandwidth_allocation(pro_type, exce, os.path.join(path, 'book_new.xls'))
            return JsonResponse({'status': 'ok', 'msg': '计算完成'})

        else:
            if check_box_list[0] == '1':

                hx = list(hx.keys())

                lhx_fn(hx, pro_type)

                all_exce = glob.glob(path + "/*.xls")

                for exce in all_exce:
                    fh = open_exce(exce)
                    sheet = fh.sheet_by_name('直联')  # 打开sheet页
                    sheet1 = fh.sheet_by_name('点点路由')
                    sheet_list = fh.sheet_names()
                    nodes = list(node_dict.keys())
                    num = len(nodes)

                    book2 = copy(fh)

                    sheet_index = sheet_list.index('点点路由')
                    sh = book2.get_sheet(sheet_index)
                    #
                    # print(exce, '进行最短寻路中。。。')
                    nlist = []

                    for i in range(1, num + 1):
                        elist = []
                        for j in range(1, num + 1):

                            bandwidth = sheet.cell_value(i, j)

                            if i == j or bandwidth == '' or bandwidth == 0:
                                continue
                            else:
                                edge = Edge(nodes[i - 1], nodes[j - 1], bandwidth)

                                if nodes[i - 1] == edge.startNodeId:
                                    elist.append(edge)
                        if elist == []:
                            continue
                        else:
                            node = Node(nodes[i - 1], elist)
                        nlist.append(node)
                    graph = Graph(nlist)
                    name = os.listdir(path)

                    for na in name:
                        if na[:-4] == exce[-7:-4]:
                            n = exce[-7:-4]

                            for startNodeId in nodes:

                                for endNodeId in nodes:

                                    if startNodeId != endNodeId and (
                                            startNodeId == nodes.index(n) or endNodeId == nodes.index(n)):
                                        continue

                                    else:

                                        originNodeId, destNodeId, route, weight = graph.dijkstra(startNodeId, endNodeId)

                                    for j in range(1, num + 1):
                                        for z in range(1, num + 1):
                                            StoE = sheet1.cell((j - 1) * num + z, 0).value
                                            if originNodeId == destNodeId:
                                                sh.write((j - 1) * num + z, 1, 'NULL')
                                                sh.write((j - 1) * num + z, 2, 'NULL')
                                                sh.write((j - 1) * num + z, 3, 'NULL')
                                                sh.write((j - 1) * num + z, 4, 'NULL')
                                            else:

                                                if originNodeId + "-" + destNodeId == StoE:
                                                    sh.write((j - 1) * num + z, 1, len(route))
                                                    sh.write((j - 1) * num + z, 2, weight)

                                                    for p in route:
                                                        pa = p + [destNodeId]
                                                        sh.write((j - 1) * num + z, 3, len(pa) - 1)
                                                        sh.write((j - 1) * num + z, 4 + route.index(p), '-'.join(pa))
                        logger.info(na, '计算中。。。。。')
                    if os.path.exists(exce):
                        book2.save(os.path.join(path, 'book_new.xls'))  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
                        os.remove(exce)
                        os.rename(os.path.join(path, 'book_new.xls'), exce)
                    else:
                        book2.save(exce)
                    bandwidth_allocation(pro_type, exce, os.path.join(path, 'book_new.xls'))
                return JsonResponse({'status': 'ok', 'msg': '计算完成'})

            elif check_box_list[0] == '2':

                lhx = list(lhx.keys())

                lhx_fn(lhx, pro_type)

                all_exce = glob.glob(path + "/*.xls")

                for exce in all_exce:
                    fh = open_exce(exce, )
                    sheet = fh.sheet_by_name('直联')  # 打开sheet页
                    sheet1 = fh.sheet_by_name('点点路由')
                    sheet_list = fh.sheet_names()
                    nodes = list(node_dict.keys())
                    num = len(nodes)

                    book2 = copy(fh)

                    sheet_index = sheet_list.index('点点路由')
                    sh = book2.get_sheet(sheet_index)

                    logger.info(exce, '进行最短寻路中。。。')
                    nlist = []

                    for i in range(1, num + 1):
                        elist = []
                        for j in range(1, num + 1):

                            bandwidth = sheet.cell_value(i, j)

                            if i == j or bandwidth == '' or bandwidth == 0:
                                continue
                            else:
                                edge = Edge(nodes[i - 1], nodes[j - 1], bandwidth)

                                if nodes[i - 1] == edge.startNodeId:
                                    elist.append(edge)
                        if elist == []:
                            continue
                        else:
                            node = Node(nodes[i - 1], elist)
                        nlist.append(node)
                    graph = Graph(nlist)
                    name = os.listdir(path)

                    for na in name:
                        if na[:-4] == exce[-7:-4]:
                            na = exce[-7:-4]

                            for i in nodes:

                                startNodeId = i

                                for j in nodes:
                                    endNodeId = j
                                    if startNodeId == na or endNodeId == na:
                                        continue

                                    else:

                                        originNodeId, destNodeId, route, weight = graph.dijkstra(startNodeId, endNodeId)

                                    for j in range(1, num + 1):
                                        for z in range(1, num + 1):
                                            StoE = sheet1.cell((j - 1) * num + z, 0).value
                                            if originNodeId == destNodeId:
                                                sh.write((j - 1) * num + z, 1, 'NULL')
                                                sh.write((j - 1) * num + z, 2, 'NULL')
                                                sh.write((j - 1) * num + z, 3, 'NULL')
                                                sh.write((j - 1) * num + z, 4, 'NULL')

                                            if StoE == originNodeId + "-" + destNodeId:
                                                sh.write((j - 1) * num + z, 1, len(route))
                                                sh.write((j - 1) * num + z, 2, weight)

                                                for p in route:
                                                    pa = p + [destNodeId]
                                                    sh.write((j - 1) * num + z, 3, len(pa) - 1)
                                                    sh.write((j - 1) * num + z, 4 + route.index(p), '-'.join(pa))

                    if os.path.exists(exce):
                        book2.save(os.path.join(path, 'book_new.xls'))  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
                        os.remove(exce)
                        os.rename(os.path.join(path, 'book_new.xls'), exce)
                    else:
                        book2.save(exce)
                    bandwidth_allocation(pro_type, exce, os.path.join(path, 'book_new.xls'))
                return JsonResponse({'status': 'ok', 'msg': '计算完成'})

            elif check_box_list[0] == '3':

                nodelist = list(node_dict.keys())

                wb = open_exce(os.path.join(os.path.abspath('excel/cmnetwin'), 'CMNET网_直联.xls'))
                sheet1 = wb.sheet_by_index(0)

                num = len(nodelist)

                dict = {}
                for i in range(1, num + 1):
                    for j in range(1, num + 1):
                        ban = sheet1.cell(i, j).value

                        if ban != '' and ban != 0:
                            if (nodelist[i - 1] != nodelist[j - 1]):
                                dict[nodelist[i - 1] + '-' + nodelist[j - 1]] = ban

                max_fn(nodelist, pro_type)

                all_exce = glob.glob(path + '/*.xls')

                # 得到要合并的所有exce表格数据
                if (all_exce == 0):
                    logger.info("该目录下无.xlsx文件！请检查您输入的目录是否有误！")
                    os.system('pause')
                    exit()

                for exce in all_exce:
                    fh = open_exce(exce)

                    sheet = fh.sheet_by_name('直联')  # 打开sheet页
                    sheet1 = fh.sheet_by_name('点点路由')
                    sheet_list = fh.sheet_names()
                    nodes = sheet.row_values(0)[1:]

                    num = len(nodes)

                    book2 = copy(fh)

                    sheet_index = sheet_list.index('点点路由')
                    sh = book2.get_sheet(sheet_index)
                    # print(exce, '最短寻路进行中。。。')

                    nlist = []

                    for i in range(1, num + 1):
                        elist = []
                        for j in range(1, num + 1):

                            bandwidth = sheet.cell_value(i, j)

                            if i == j or bandwidth == '' or bandwidth == 0:
                                continue
                            else:
                                edge = Edge(nodes[i - 1], nodes[j - 1], bandwidth)

                                if nodes[i - 1] == edge.startNodeId:
                                    elist.append(edge)

                            node = Node(nodes[i - 1], elist)
                        nlist.append(node)

                    graph = Graph(nlist)

                    for startNodeId in nodes:

                        for endNodeId in nodes:

                            if startNodeId == endNodeId:
                                continue
                            else:

                                originNodeId, destNodeId, route, weight = graph.dijkstra(startNodeId, endNodeId)

                            for j in range(1, num + 1):
                                for z in range(1, num + 1):
                                    StoE = sheet1.cell((j - 1) * num + z, 0).value
                                    if originNodeId == destNodeId:
                                        sh.write((j - 1) * num + z, 1, 'NULL')
                                        sh.write((j - 1) * num + z, 2, 'NULL')
                                        sh.write((j - 1) * num + z, 3, 'NULL')
                                        sh.write((j - 1) * num + z, 4, 'NULL')

                                    if StoE == originNodeId + "-" + destNodeId:
                                        sh.write((j - 1) * num + z, 1, len(route))
                                        sh.write((j - 1) * num + z, 2, weight)

                                        for p in route:
                                            path = p + [destNodeId]
                                            sh.write((j - 1) * num + z, 4 + route.index(p), '-'.join(path))
                                            sh.write((j - 1) * num + z, 3, len(path) - 1)
                                print(exce, originNodeId, destNodeId, '计算中。。。。。')

                    if os.path.exists(exce):
                        book2.save(os.path.join(path, 'book_new.xls'))  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
                        os.remove(exce)
                        os.rename(os.path.join(path, 'book_new.xls'), exce)
                    else:
                        book2.save(exce)
                    bandwidth_allocation(pro_type, exce, os.path.join(path, 'book_new.xls'))
                return JsonResponse({'status': 'ok', 'msg': '计算完成'})
            else:
                return JsonResponse({'status': 'ok', 'msg': '计算中出现错误'})
    else:
        return JsonResponse({'status': 'ok', 'msg': '计算中出现错误'})


# 链路故障（一条或多条）
def link_fail(request):
    if request.method == 'GET':
        link = request.GET.get('arr')
        link = eval(link)
        pro_type = '2'

        pm = ProjectInfo.objects.get(id=2).pm

        node_dict, hx, lhx, hj = find_node(pm)

        metric_path = settings.MEDIA_ROOT + 'cmnetwin/metric/'
        path = os.path.abspath('excel/cmnetwin/min')
        if not os.path.exists(metric_path):
            return JsonResponse({'status': 'err', 'msg': '请上传metric表'})
        if not os.path.exists(path):
            os.mkdir(path)

        all_exce = glob.glob(metric_path + "/*.xls")
        for exce in all_exce:
            exc = open_exce(exce)
            metric_sh = exc.sheet_by_name('Metric')
            sheet_list = exc.sheet_names()
            nodes = list(node_dict.keys())

            num = len(nodes)

            for l in link:
                new_exce = copy(exc)
                sheet_index = sheet_list.index('Metric')
                sh = new_exce.get_sheet(sheet_index)
                sh2 = new_exce.add_sheet('直联', cell_overwrite_ok=True)
                sh3 = new_exce.add_sheet('点点路由', cell_overwrite_ok=True)
                st = l['st']
                en = l['en']

                sh.write(nodes.index(st) + 1, nodes.index(en) + 1, '', gz_style)
                for i in range(1, num + 1):
                    start_node = metric_sh.cell_value(i, 0)
                    sh2.write(i, 0, start_node)
                    sh2.write(0, i, start_node)
                    for j in range(1, num + 1):
                        end_node = metric_sh.cell_value(0, j)
                        band = metric_sh.cell(i, j).value

                        sh2.write(i, j, band)
                        sh2.write(j, i, band)
                        sh3.write((i - 1) * num + j, 0, start_node + '-' + end_node)
                        if i == j:
                            sh3.write((i - 1) * num + j, 1, 'NULL')
                            sh3.write((i - 1) * num + j, 2, 'NULL')
                            sh3.write((i - 1) * num + j, 3, 'NULL')
                            sh3.write((i - 1) * num + j, 4, 'NULL')
                sh2.write(nodes.index(st) + 1, nodes.index(en) + 1, '', gz_style)
                sh2.write(nodes.index(en) + 1, nodes.index(st) + 1, '', gz_style)

                sh3.write(0, 0, '源节点-目的节点')
                sh3.write(0, 1, '同权路由数量')
                sh3.write(0, 2, '度量值')
                sh3.write(0, 3, '路由跳数')
                sh3.write(0, 4, '路由')
                new_exce.save(os.path.join(path, st + '-' + en + '.xls'))

        path2 = os.path.join(path, 'book_new.xls')
        all_link = glob.glob(path + "/*.xls")

        for exce in all_link:
            exc = open_exce(exce)

            sheet_list = exc.sheet_names()
            sheet = exc.sheet_by_name('直联')
            sheet1 = exc.sheet_by_name('点点路由')
            book2 = copy(exc)

            sheet_index = sheet_list.index('点点路由')
            sh = book2.get_sheet(sheet_index)

            nodes = list(node_dict.keys())

            num = len(nodes)

            nlist = []
            for i in range(1, num + 1):
                elist = []
                for j in range(1, num + 1):

                    bandwidth = sheet.cell(i, j).value

                    if i == j or bandwidth == '' or bandwidth == 0:
                        continue
                    else:
                        edge = Edge(nodes[i - 1], nodes[j - 1], bandwidth)

                        if nodes[i - 1] == edge.startNodeId:
                            elist.append(edge)

                        node = Node(nodes[i - 1], elist)

                    nlist.append(node)

            graph = Graph(nlist)
            for startNodeId in nodes:

                for endNodeId in nodes:

                    if startNodeId == endNodeId:
                        continue
                    else:

                        originNodeId, destNodeId, route, weight = graph.dijkstra(startNodeId, endNodeId)

                    for j in range(1, num + 1):
                        for z in range(1, num + 1):
                            StoE = sheet1.cell((j - 1) * num + z, 0).value

                            if originNodeId == destNodeId:
                                sh.write((j - 1) * num + z, 1, 'NULL')
                                sh.write((j - 1) * num + z, 2, 'NULL')
                                sh.write((j - 1) * num + z, 3, 'NULL')
                                sh.write((j - 1) * num + z, 4, 'NULL')

                            if originNodeId + "-" + destNodeId == StoE:
                                sh.write((j - 1) * num + z, 1, len(route))
                                sh.write((j - 1) * num + z, 2, weight)

                                for p in route:
                                    pa = p + [destNodeId]
                                    sh.write((j - 1) * num + z, 3, len(pa) - 1)
                                    sh.write((j - 1) * num + z, 4 + route.index(p), '-'.join(pa))

                    print(originNodeId, destNodeId, '计算中。。。。。')

            if os.path.exists(exce):
                os.remove(exce)
                book2.save(path2)

                os.rename(path2, exce)
            else:
                book2.save(exce)
            bandwidth_allocation(pro_type, exce, path2)
        return JsonResponse({'status': 'ok', 'msg': '计算完成'})
    else:
        return JsonResponse({'status': 'ok', 'msg': '计算中出现错误'})


# 搜索点点路由
def search_routefile(request):
    if request.method == 'POST':
        node_key = request.POST.get('val')

        dir_name = os.path.abspath('excel/cmnetwin/max')
        if not os.path.exists(dir_name):
            return JsonResponse({'status': 'fail', 'msg': '请先进行故障轮询计算！！'})
        else:
            file_name = os.listdir(dir_name)  # 路径下文件名称
            route_list = []
            for i in file_name:
                route_list.append(i[:-4])

            res = []
            for i in route_list:
                if node_key in i:
                    res.append(i)
            jsonArr = json.dumps(res, ensure_ascii=False)

            return JsonResponse({'status': 'ok', 'msg': jsonArr})


# 下载点点路由
def down_route(request):
    if request.method == 'POST':
        route = request.POST.get('node1')

        path = os.path.join(os.path.abspath('excel/cmnetwin/max'), route + '.xls')

        wb = open_exce(path)
        ws = wb.sheet_by_name('点点路由')
        book = xlwt.Workbook()
        sh = book.add_sheet('点点路由', cell_overwrite_ok=True)
        row = ws.nrows
        col = ws.ncols
        for i in range(row):
            for j in range(col):
                rou = ws.cell(i, j).value

                sh.write(i, j, rou)

        response = HttpResponse(content_type='application/octet-stream')
        response['Content-Disposition'] = 'attachment; filename="{0}"'.format(
            escape_uri_path('CMNET网_' + route + '.xls'))
        book.save(response)

        return response


# 中继计算（10GE，100GE，400GE）
def relay_calculation(request):
    if request.method == 'POST':
        v1 = request.POST.get('v1')
        v2 = request.POST.get('v2')
        v3 = request.POST.get('v3')
        v4 = request.POST.get('v4')
        v5 = request.POST.get('v5')
        v6 = request.POST.get('v6')
        v7 = request.POST.get('v7')
        bandwidth1 = request.POST.get('bandwidth1')
        v8 = request.POST.get('v8')
        bandwidth2 = request.POST.get('bandwidth2')
        v9 = request.POST.get('v9')
        bandwidth3 = request.POST.get('bandwidth3')
        zs1 = request.POST.get('zs1')
        zs2 = request.POST.get('zs2')
        plan = request.POST.get('plan')
        granularity = request.POST.get('granularity')
        xs = ProjectInfo.objects.get(id=2).xs
        pm = ProjectInfo.objects.get(id=2).pm

        node_dict, hx, lhx, hj = find_node(pm)

        hx = list(hx.keys())
        lhx = list(lhx.keys())
        hj = list(hj.keys())

        path1 = os.path.join(os.path.abspath('excel/cmnetwin'), 'CMNET网工程计算.xls')
        path2 = os.path.join(os.path.abspath('excel/cmnetwin'), 'newCMNET网工程计算.xls')

        v1 = int(v1) / 100 / xs
        v2 = int(v2) / 100 / xs
        v3 = int(v3) / 100 / xs
        v4 = int(v4) / 100 / xs
        v5 = int(v5) / 100 / xs
        v6 = int(v6) / 100 / xs

        if granularity == '100GE':
            band_100GE(node_dict, path1, path2, hx, lhx, hj, zs1, zs2, v1, v2, v3, v4, v5, v6, v7, bandwidth1, v8,
                       bandwidth2, v9,
                       bandwidth3)
        elif granularity == '10GE':
            band_10GE(node_dict, path1, path2, hx, lhx, hj, zs1, zs2, v1, v2, v3, v4, v5, v6, v7, bandwidth1, v8,
                      bandwidth2, v9,
                      bandwidth3)
        elif granularity == '10GPOS':
            band_10GE(node_dict, path1, path2, hx, lhx, hj, zs1, zs2, v1, v2, v3, v4, v5, v6, v7, bandwidth1, v8,
                      bandwidth2, v9,
                      bandwidth3)
        elif granularity == '400GE':
            band_400GE(node_dict, path1, path2, hx, lhx, hj, zs1, zs2, v1, v2, v3, v4, v5, v6, v7, bandwidth1, v8,
                       bandwidth2, v9,
                       bandwidth3)

        plan_path = os.path.abspath('excel/cmnetwin/plan')
        p1 = os.path.abspath('excel/cmnetwin')
        if not os.path.exists(plan_path):
            os.makedirs(plan_path)

        if plan == '1':
            wb = open_exce(os.path.join(p1, 'CMNET网工程计算.xls'))
            b2 = copy(wb)

            b2.save(os.path.join(plan_path, '方案1.xls'))
        else:
            logger.info('不做方案对比')

        return JsonResponse({'status': 'ok', 'msg': '计算完成'})
    else:
        return JsonResponse({'status': 'ok', 'msg': '计算失败'})


# 中继比较
def relay_compare(request):
    if request.method == 'POST':
        relayFile = request.FILES.get('relayfile')

        path = settings.MEDIA_ROOT + 'cmnetwin/relay/'  # 上传文件的保存路径，可以自己指定任意的路径
        if not relayFile:
            # 有数据为空
            return JsonResponse({'status': 'err', 'msg': '参数不能为空!'})
        if not os.path.exists(path):
            os.makedirs(path)
            with open(os.path.join(path, relayFile.name), 'wb+') as f:
                for chunk in relayFile.chunks():
                    f.write(chunk)
        else:
            try:
                with open(os.path.join(path, relayFile.name), 'wb+') as f:
                    for chunk in relayFile.chunks():
                        f.write(chunk)
            except PermissionError:
                return JsonResponse({'status': 'err', 'msg': '同名文件已打开，请关闭文件重新导入'})
        try:
            relay_direction()  # 中继方向
            pm = ProjectInfo.objects.get(id=2).pm

            node_dict, hx, lhx, hj = find_node(pm)

            excel = xlrd.open_workbook(os.path.join(path, relayFile.name))  # 上期中继
            sheetcount = len(excel.sheet_names())
            if sheetcount == 2:

                sheet = excel.sheet_by_index(0)
                sheet2 = excel.sheet_by_index(1)
            elif sheetcount == 3:
                sheet = excel.sheet_by_index(0)
                sheet2 = excel.sheet_by_index(1)
                sheet3 = excel.sheet_by_index(2)
            elif sheetcount == 1:
                sheet = excel.sheet_by_index(0)

            nodes = list(node_dict.keys())
            num = len(nodes)

            path1 = os.path.join(os.path.abspath('excel/cmnetwin'), 'CMNET网工程计算.xls')
            path2 = os.path.join(os.path.abspath('excel/cmnetwin'), 'newCMNET网工程计算.xls')

            wb = open_exce(path1)
            sh1 = wb.sheet_by_name('中继带宽T2')

            # # ===========各省中继方向统计=====================================================

            book2 = copy(wb)
            sheets = wb.sheet_names()

            #
            dd = {}
            d1 = {}
            for i in range(1, num + 1):
                for j in range(1, num + 1):
                    relay = sheet.cell(i, j).value
                    dd[nodes[i - 1] + '~' + nodes[j - 1]] = relay
                    if relay != 0 and relay != '':
                        d1[nodes[i - 1] + '~' + nodes[j - 1]] = relay
            d2 = {}
            for i in range(1, num + 1):
                for j in range(1, num + 1):
                    relay = sheet2.cell(i, j).value
                    if relay != 0 and relay != '':
                        d2[nodes[i - 1] + '~' + nodes[j - 1]] = relay
            d3 = {}
            for i in range(1, num + 1):
                for j in range(1, num + 1):
                    relay = sh1.cell(i, j).value
                    if relay != 0 and relay != '':
                        d3[nodes[i - 1] + '~' + nodes[j - 1]] = relay

            relaydic = dict(list(d1.items()) + list(d2.items()))

            sorted_dict = list(map(lambda x: {x: relaydic[x]} if x in relaydic.keys() else 0, list(dd.keys())))
            res = {}  # 去除不存在的链路
            for i in sorted_dict:
                if i == 0:
                    continue
                else:
                    res.update(i)

            for k, v in d3.items():
                if k in list(res.keys()):
                    continue
                else:
                    res[k] = v
            # print(res, len(res))

            if '分省电路表' in sheets:
                index = sheets.index('分省电路表')
                sh = book2.get_sheet(index)

            else:
                sh = add_sheet(book2,'分省电路表')
                sec_col = sh.col(1)
                sec_col.width = 256 * 20

                sh.write_merge(0, 1, 0, 0, '序号', style)  # 以合并单元格形式写入数据，即将数据写入
                sh.write_merge(0, 1, 1, 1, '电路局向', style)
                sh.write_merge(0, 0, 2, 5, '往期到达', style)
                sh.write_merge(0, 0, 6, 9, '本期到达', style)
                sh.write_merge(0, 0, 10, 13, '本期新增', style)
                sh.write_merge(0, 0, 14, 17, '本期撤销', style)
                # =====================往期到达==============================

                sh.write_merge(1, 1, 2, 2, u'100GE', style1)
                sh.write_merge(1, 1, 3, 3, u'10GPOS', style1)
                sh.write_merge(1, 1, 4, 4, u'10GE', style1)
                sh.write_merge(1, 1, 5, 5, u'400GE', style1)

                # =====================本期到达==============================

                sh.write_merge(1, 1, 6, 6, u'100GE', style1)
                sh.write_merge(1, 1, 7, 7, u'10GPOS', style1)
                sh.write_merge(1, 1, 8, 8, u'10GE', style1)
                sh.write_merge(1, 1, 9, 9, u'400GE', style1)

                # =====================本期新增到达==============================

                sh.write_merge(1, 1, 10, 10, u'100GE', style1)
                sh.write_merge(1, 1, 11, 11, u'10GPOS', style1)
                sh.write_merge(1, 1, 12, 12, u'10GE', style1)
                sh.write_merge(1, 1, 13, 13, u'400GE', style1)

                # =====================本期撤销到达==============================

                sh.write_merge(1, 1, 14, 14, u'100GE', style1)
                sh.write_merge(1, 1, 15, 15, u'10GPOS', style1)
                sh.write_merge(1, 1, 16, 16, u'10GE', style1)
                sh.write_merge(1, 1, 17, 17, u'400GE', style1)

            for i in range(2, len(res) + 3):

                    if i != len(res) + 2:
                        sh.write(i, 0, i - 1, style2)
                        for j in range(2, 15):
                            sh.write(i, j - 1, '', style2)
                    else:

                        sh.write(i, 0, '', he_style3)
                        sh.write(i, 1, '合计', he_style3)
                        sh.write(i, 2, Formula('SUM(C3:C%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 3, Formula('SUM(D3:D%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 4, Formula('SUM(E3:E%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 5, Formula('SUM(F3:F%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 6, Formula('SUM(G3:G%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 7, Formula('SUM(H3:H%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 8, Formula('SUM(I3:I%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 9, Formula('SUM(J3:J%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 10, Formula('SUM(K3:K%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 11, Formula('SUM(L3:L%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 12, Formula('SUM(M3:M%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 13, Formula('SUM(N3:N%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 14, Formula('SUM(O3:O%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 15, Formula('SUM(P3:P%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 16, Formula('SUM(Q3:Q%s)' % str(len(res) + 2)), he_style3)
                        sh.write(i, 17, Formula('SUM(R3:R%s)' % str(len(res) + 2)), he_style3)

            for k, v in res.items():

                sh.write(list(res.keys()).index(k) + 2, 1, k, style3)

                if isinstance(v, float):
                    if k in list(d1.keys()):

                        sh.write(list(res.keys()).index(k) + 2, 2, d1[k], style3)
                    elif k in list(d2.keys()):
                        sh.write(list(res.keys()).index(k) + 2, 4, d2[k], style3)

                if k in list(d3.keys()):

                    n = d3[k].split('*')

                    if 'POS' in d3[k]:
                        sh.write(list(res.keys()).index(k) + 2, 7, int(n[0]), style3)
                    elif '100GE' in d3[k]:
                        sh.write(list(res.keys()).index(k) + 2, 6, int(n[0]), style3)
                    elif '10GE' in d3[k]:
                        sh.write(list(res.keys()).index(k) + 2, 8, int(n[0]), style3)
                    elif '400GE' in d3[k]:
                        sh.write(list(res.keys()).index(k) + 2, 9, int(n[0]), style3)
                    elif isinstance(d3[k], float):
                        sh.write(list(res.keys()).index(k) + 2, 6, int(n[0]), style3)

                else:

                    n = v.split('*')

                    if 'POS' in v:
                        sh.write(list(res.keys()).index(k) + 2, 7, int(n[0]), style3)
                    elif '100GE' in v:
                        sh.write(list(res.keys()).index(k) + 2, 6, int(n[0]), style3)
                    elif '10GE' in v:
                        sh.write(list(res.keys()).index(k) + 2, 8, int(n[0]), style3)
                    elif '400GE' in d3[k]:
                        sh.write(list(res.keys()).index(k) + 2, 9, int(n[0]), style3)
                    elif isinstance(v, float):
                        sh.write(list(res.keys()).index(k) + 2, 6, int(n[0]), style3)

                sh.write(list(res.keys()).index(k) + 2, 10, Formula('IF((G%s-C%s)<0,0,(G%s-C%s))' % (
                str(len(res) + 2), str(len(res) + 2), str(len(res) + 2), str(len(res) + 2))), style3)
                sh.write(list(res.keys()).index(k) + 2, 11, Formula('IF((H%s-D%s)<0,0,(H%s-D%s))' % (
                str(len(res) + 2), str(len(res) + 2), str(len(res) + 2), str(len(res) + 2))), style3)
                sh.write(list(res.keys()).index(k) + 2, 12, Formula('IF((I%s-E%s)<0,0,(I%s-E%s))' % (
                str(len(res) + 2), str(len(res) + 2), str(len(res) + 2), str(len(res) + 2))),
                         style3)
                sh.write(list(res.keys()).index(k) + 2, 13, Formula('IF((J%s-F%s)<0,0,(J%s-F%s))' % (
                str(len(res) + 2), str(len(res) + 2), str(len(res) + 2), str(len(res) + 2))),
                         style3)
                sh.write(list(res.keys()).index(k) + 2, 14, Formula('IF((G%s-C%s)<0,-(G%s-C%s),0)' % (
                str(len(res) + 2), str(len(res) + 2), str(len(res) + 2), str(len(res) + 2))),
                         style3)
                sh.write(list(res.keys()).index(k) + 2, 15, Formula('IF((H%s-D%s)<0,-(H%s-D%s),0)' % (
                str(len(res) + 2), str(len(res) + 2), str(len(res) + 2), str(len(res) + 2))),
                         style3)
                sh.write(list(res.keys()).index(k) + 2, 16, Formula('IF((I%s-E%s)<0,-(I%s-E%s),0)' % (
                str(len(res) + 2), str(len(res) + 2), str(len(res) + 2), str(len(res) + 2))),
                         style3)
                sh.write(list(res.keys()).index(k) + 2, 17, Formula('IF((J%s-F%s)<0,-(J%s-F%s),0)' % (
                str(len(res) + 2), str(len(res) + 2), str(len(res) + 2), str(len(res) + 2))),
                         style3)

            save_exce(book2,path1,path2)

            relay_direction_statistics()
            path = os.path.abspath('excel/cmnetwin')
            flow_path = os.path.join(path, 'CMNET网工程计算.xls')
            flow_path2 = os.path.join(path, 'newCMNET网工程计算.xls')

            book = open_exce(flow_path)  # 打开文件，并且保留原格式
            sheet = book.sheet_by_name('中继方向')

            sheet_list = book.sheet_names()
            book2 = copy(book)


            pm = ProjectInfo.objects.get(id=2).pm
            data1 = CNodes_detail.objects.filter(plane1=pm, part1__in=['BB', 'BC']).exclude(program='新建').values(
                'province', 'city')
            data2 = CNodes_detail.objects.filter(plane2=pm, part2__in=['BB', 'BC']).exclude(program='新建').values(
                'province', 'city')

            olddic = {}
            newdic = {}
            province = []
            province2 = []

            for i in data1:

                if i['province'] in list(olddic.keys()):
                    olddic[i['province']] += [i['city']]
                else:
                    olddic[i['province']] = [i['city']]

                if i['province'] in province:
                    continue
                else:
                    province.append(i['province'])
                if i['province'] in province2:
                    continue
                else:
                    province2.append(i['province'][:2])

            for i in data2:

                if i['province'] in list(newdic.keys()):
                    newdic[i['province']] += [i['city']]
                else:
                    newdic[i['province']] = [i['city']]
            # print('old',olddic)
            # print('new',newdic)

            node_dict, hx, lhx, hj = find_node(pm)

            hxlist = list(hx.keys()) + list(lhx.keys())

            hjlist = list(hj.keys())


            # # ===========各省中继方向统计=====================================================
            if '分省中继方向统计' in sheet_list:
                innum = sheet_list.index('分省中继方向统计')
                relay_sh = book2.get_sheet(innum)

            else:
                relay_sh = add_sheet(book2,'分省中继方向统计')
                relay_sh.write_merge(0, 1, 0, 0, '方向数量', style)  # 以合并单元格形式写入数据，即将数据写入
                relay_sh.write_merge(0, 0, 1, 2, '上期到达', style)
                relay_sh.write_merge(1, 1, 1, 1, '至核心', style)
                relay_sh.write_merge(1, 1, 2, 2, '至汇聚', style)
                relay_sh.write_merge(0, 0, 3, 4, '本期到达', style)
                relay_sh.write_merge(1, 1, 3, 3, '至核心', style)
                relay_sh.write_merge(1, 1, 4, 4, '至汇聚', style)
                relay_sh.write_merge(0, 0, 5, 6, '本期新增', style)
                relay_sh.write_merge(1, 1, 5, 5, '至核心', style)
                relay_sh.write_merge(1, 1, 6, 6, '至汇聚', style)


            for i in range(2, len(province) + 2):
                relay_sh.write(i, 0, province[i - 2], style)
                relay_sh.write(i, 5, Formula('IF((D%s-B%s)<0,0,(D%s-B%s))' % (
                    str(len(province2) + 2), str(len(province2) + 2), str(len(province2) + 2),
                    str(len(province2) + 2))),
                               style)
                relay_sh.write(i, 6, Formula('IF((E%s-C%s)<0,0,(E%s-C%s))' % (
                    str(len(province2) + 2), str(len(province2) + 2), str(len(province2) + 2),
                    str(len(province2) + 2))),
                               style)
            # print(relaydic)
            prodic = {}
            newprodic = {}
            for p in province2:
                o = []
                n = []

                for k,v in relaydic.items():# 上期工程中继方向数据格式化
                    remove_digits = str.maketrans('', '', digits)
                    res = k.translate(remove_digits)
                    link = res.split('~')
                    if p in link:
                        o.append(k)

                for k,v in d3.items(): #本期工程中继方向数据格式化
                    remove_digits = str.maketrans('', '', digits)
                    res = k.translate(remove_digits)
                    link2 = res.split('~')
                    if p in link2:
                        n.append(k)
                prodic[p] = o
                newprodic[p] = n

            relaylink = {}
            for k,v in prodic.items():
                num = 0
                num2 = 0

                for i in v:
                    l = i.split('~')
                    if l[0] in hxlist or l[1] in hxlist:
                        num += 1
                    if l[0] in hjlist or l[1] in hjlist:
                        num2 += 1
                relaylink[k]=[num,num2]

            newrelaylink = {}
            for k, v in newprodic.items():
                num = 0
                num2 = 0

                for i in v:
                    l = i.split('~')
                    if l[0] in hxlist or l[1] in hxlist:
                        num += 1
                    if l[0] in hjlist or l[1] in hjlist:
                        num2 += 1
                newrelaylink[k] = [num, num2]

            # print(relaylink)
            for k,v in relaylink.items():
                relay_sh.write(province2.index(k)+2,1,v[0],style)
                relay_sh.write(province2.index(k) + 2, 2, v[1], style)

            for k,v in newrelaylink.items():
                relay_sh.write(province2.index(k)+2,3,v[0],style)
                relay_sh.write(province2.index(k) + 2, 4, v[1], style)

            save_exce(book2,flow_path,flow_path2)


            return JsonResponse({'status': 'ok', 'msg': '计算完成！！'})

        except PermissionError:
            return JsonResponse({'status': 'err', 'msg': '你有可能已经打开了这个文件，关闭这个文件即可！！'})

# 中继电路回显
def show_trunkcircuit(request):
    try:
        if request.method == "GET":
            path = os.path.join(os.path.abspath('excel/cmnetwin'), 'CMNET网工程计算.xls')
            wb = open_exce(path)
            sheets = wb.sheet_names()
            if '分省电路表' in sheets:
                index = sheets.index('分省电路表')
                ws = wb.sheet_by_index(index)
            else:
                return JsonResponse({'status': 'err', 'msg': '请先导入往期中继计算！！'})

            rows = ws.nrows
            cols = ws.ncols
            # 创建一个数组用来存储excel中的数据
            p = []
            for i in range(2, rows):
                d = {}
                for j in range(0, cols):

                    d['num'] = ws.cell(i, 0).value #序号
                    d['trunkcircuit'] = ws.cell(i, 1).value # 中继电路
                    d['old100GE'] = ws.cell(i, 2).value  # 上期100GE
                    d['old10GPOS'] = ws.cell(i, 3).value
                    d['old10GE'] = ws.cell(i, 4).value
                    d['old400GE'] = ws.cell(i, 5).value
                    d['new100GE'] = ws.cell(i, 6).value  # 本期100GE
                    d['new10GPOS'] = ws.cell(i, 7).value
                    d['new10GE'] = ws.cell(i, 8).value
                    d['new400GE'] = ws.cell(i, 9).value
                    d['increase100GE'] = ws.cell(i, 10).value  # 新增100GE
                    d['increase10GPOS'] = ws.cell(i, 11).value
                    d['increase10GE'] = ws.cell(i, 12).value
                    d['increase400GE'] = ws.cell(i, 13).value
                    d['had100GE'] = ws.cell(i, 14).value  # 下撤100GE
                    d['had10GPOS'] = ws.cell(i, 15).value
                    d['had10GE'] = ws.cell(i, 16).value
                    d['had400GE'] = ws.cell(i, 17).value

                ap = []
                for k, v in d.items():
                    if isinstance(v, float):  # excel中的值默认是float,需要进行判断处理，通过'"%s":%d'，'"%s":"%s"'格式化数组
                        ap.append('"%s":%d' % (k, v))
                    else:
                        ap.append('"%s":"%s"' % (k, v))
                s = '{%s}' % (','.join(ap))  # 继续格式化
                p.append(s)
            t = '[%s]' % (','.join(p))  # 格式化

            # print(json_data)
            # print(type(json_data)) #str
            json_data = json.loads(t)

            return JsonResponse(json_data, safe=False)
        else:
            return JsonResponse({'status': 'err', 'msg': '非法请求方式！！'})
    except Exception as e:
        logger.error(e)
        return JsonResponse({'status': 'err', 'msg': e.args})


# 中继方向
def relay_direction():
    path = os.path.abspath('excel/cmnetwin')
    flow_path = os.path.join(path, 'CMNET网工程计算.xls')
    flow_path2 = os.path.join(path, 'newCMNET网工程计算.xls')

    book = open_exce(flow_path)  # 打开文件，并且保留原格式
    sheet = book.sheet_by_name('直联')

    sheet1 = book.sheet_by_name('中继带宽100GE')  # 打开sheet页

    pm = ProjectInfo.objects.get(id=2).pm

    node_dict, hx, lhx, hj = find_node(pm)
    nodes = list(node_dict.keys())
    hx = list(hx.keys())
    lhx = list(lhx.keys())
    hj = list(hj.keys())

    book2 = copy(book)
    sheet_list = book.sheet_names()

    if '中继方向' in sheet_list:
        sheet_index = sheet_list.index('中继方向')
        sh = book2.get_sheet(sheet_index)
        for i in range(1, len(nodes) + 1):
            node = sheet.cell(i, 0).value
            # print(node)

            if node in hx:
                sh.write(i + 1, 0, 'BB-c')
                sh.write(0, i + 1, 'BB-c')
            elif node in lhx:
                sh.write(i + 1, 0, 'BB-a')
                sh.write(0, i + 1, 'BB-a')
            elif node in hj:
                sh.write(i + 1, 0, 'BC')
                sh.write(0, i + 1, 'BC')

            sh.write(i + 1, 1, node)
            sh.write(1, i + 1, node)
            for j in range(1, len(nodes) + 1):
                flow = sheet1.cell_value(i, j)

                if flow != '':
                    sh.write(i + 1, j + 1, 1)

    else:
        worksheet = book2.add_sheet('中继方向', cell_overwrite_ok=True)
        worksheet.write(0, 0, '角色')
        worksheet.write(1, 1, '节点')

        for i in range(1, len(nodes) + 1):
            node = sheet.cell(i, 0).value
            # print(node)
            if node in hx:
                worksheet.write(i + 1, 0, 'BB-c')
                worksheet.write(0, i + 1, 'BB-c')
            elif node in lhx:
                worksheet.write(i + 1, 0, 'BB-a')
                worksheet.write(0, i + 1, 'BB-a')
            elif node in hj:
                worksheet.write(i + 1, 0, 'BC')
                worksheet.write(0, i + 1, 'BC')

            worksheet.write(i + 1, 1, node)
            worksheet.write(1, i + 1, node)
            for j in range(1, len(nodes) + 1):
                flow = sheet1.cell_value(i, j)
                # print(flow)
                if flow != '':
                    worksheet.write(i + 1, j + 1, 1)

    if os.path.exists(flow_path):
        book2.save(flow_path2)  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
        os.remove(flow_path)
        os.rename(flow_path2, flow_path)
    else:
        book2.save(flow_path)


# 中继方向统计
def relay_direction_statistics():
    path = os.path.abspath('excel/cmnetwin')
    flow_path = os.path.join(path, 'CMNET网工程计算.xls')
    flow_path2 = os.path.join(path, 'newCMNET网工程计算.xls')
    book = open_exce(flow_path)  # 打开文件，并且保留原格式
    sheet = book.sheet_by_name('中继方向')

    sheet_list = book.sheet_names()
    book2 = copy(book)

    pm = ProjectInfo.objects.get(id=2).pm
    proname = ProjectInfo.objects.get(id=2).name

    node_dict, hx, lhx, hj = find_node(pm)

    hxlist = list(hx.keys())
    lhxlist = list(lhx.keys())
    hjlist = list(hj.keys())
    nodes = list(node_dict.keys())

    style = easyxf('font: bold on,name 宋体,height 240;'
                   'align: wrap on, vert centre, horiz center;'
                   'borders:left 1,right 1,top 1,bottom 1')
    style1 = easyxf('font: name 宋体,height 220;'
                    'align: wrap on, vert centre, horiz center;'
                    'borders:left 1,right 1,top 1,bottom 1')
    st = easyxf('pattern: pattern solid ,fore_colour gray_ega;'
                'align: wrap on, vert centre, horiz center;'
                'borders:left 1,right 1,top 1,bottom 1')
    # rows = sheet.ncols
    # print(rows)#64
    hx = 0
    lhx = 0
    hj = 0
    for i in range(sheet.nrows):

        if sheet.cell_value(i, 0) == 'BB-c':

            for val in sheet.row_values(i, 2, sheet.ncols):
                if val == '':
                    val = 0
                hx = hx + val

        if sheet.cell_value(i, 0) == 'BB-a':

            for val in sheet.row_values(i, 2, sheet.ncols):
                if val == '':
                    val = 0
                lhx = lhx + val

        if sheet.cell_value(i, 0) == 'BC':
            for val in sheet.row_values(i, 2, sheet.ncols):
                if val == '':
                    val = 0
                hj = hj + val

    if '中继方向统计表' in sheet_list:

        sheet_index = sheet_list.index('中继方向统计表')
        worksheet = book2.get_sheet(sheet_index)
        for i in range(5):
            for j in range(7):
                # sheet.col(j).width = 256 * 20
                worksheet.write(0, 0, proname, style)
                worksheet.write(0, 1, 'BB1之间', style)
                worksheet.write(0, 2, 'BB之间', style)
                worksheet.write(0, 3, '全部', style)
                worksheet.write(0, 4, '除外BB1之间', style)
                worksheet.write(0, 5, '除外BB之间', style)
                worksheet.write(0, 6, 'BC之间', style)
                worksheet.write(1, 0, '局向数量', style1)
                worksheet.write(2, 0, '中继方向数量', style1)
                worksheet.write(3, 0, '核心数量', style1)
                worksheet.write(4, 0, '平均方向数量', style1)

                if i != 0:
                    worksheet.write(i, j, '', style1)

                if (i == 3 and j == 6) or (i == 4 and j == 6):
                    worksheet.write(3, 6, '——', st)
                    worksheet.write(4, 6, '——', st)
                worksheet.write(1, 1, int(hx), style1)
                worksheet.write(1, 2, int(lhx), style1)
                worksheet.write(1, 3, int(hx + lhx + hj), style1)
                worksheet.write(1, 4, Formula('D2-B2'), style1)
                worksheet.write(1, 5, Formula('D2-C2'), style1)
                worksheet.write(1, 6, Formula('D2-B2-C2'), style1)
                worksheet.write(2, 1, int(hx) / 2, style1)
                worksheet.write(2, 2, int(lhx) / 2, style1)
                worksheet.write(2, 3, int(hj) / 2, style1)
                worksheet.write(2, 4, Formula('D3-B3'), style1)
                worksheet.write(2, 5, Formula('D3-C3'), style1)
                worksheet.write(2, 6, Formula('D3-B3-C3'), style1)
                worksheet.write(3, 1, len(hxlist) // 2, style1)
                worksheet.write(3, 2, (len(hxlist) + len(lhxlist)) // 2, style1)
                worksheet.write(3, 3, (len(hxlist) + len(lhxlist)) // 2, style1)
                worksheet.write(3, 4, 26 - len(hxlist), style1)
                worksheet.write(3, 5, 26 - len(hxlist), style1)
                worksheet.write(4, 1, Formula('B3/B4'), style1)
                worksheet.write(4, 2, Formula('C3/C4'), style1)
                worksheet.write(4, 3, Formula('D3/D4'), style1)
                worksheet.write(4, 4, Formula('E3/E4'), style1)
                worksheet.write(4, 5, Formula('F3/F4'), style1)

    else:
        worksheet = book2.add_sheet('中继方向统计表', cell_overwrite_ok=True)
        for i in range(5):
            for j in range(7):
                # sheet.col(j).width = 256 * 20
                worksheet.write(0, 0, proname, style)
                worksheet.write(0, 1, 'BB1之间', style)
                worksheet.write(0, 2, 'BB之间', style)
                worksheet.write(0, 3, '全部', style)
                worksheet.write(0, 4, '除外BB1之间', style)
                worksheet.write(0, 5, '除外BB之间', style)
                worksheet.write(0, 6, 'BC之间', style)
                worksheet.write(1, 0, '局向数量', style1)
                worksheet.write(2, 0, '中继方向数量', style1)
                worksheet.write(3, 0, '核心数量', style1)
                worksheet.write(4, 0, '平均方向数量', style1)

                if i != 0:
                    worksheet.write(i, j, '', style1)

                if (i == 3 and j == 6) or (i == 4 and j == 6):
                    worksheet.write(3, 6, '——', st)
                    worksheet.write(4, 6, '——', st)
                worksheet.write(1, 1, int(hx), style1)
                worksheet.write(1, 2, int(lhx), style1)
                worksheet.write(1, 3, int(hx + lhx + hj), style1)
                worksheet.write(1, 4, Formula('D2-B2'), style1)
                worksheet.write(1, 5, Formula('D2-C2'), style1)
                worksheet.write(1, 6, int(hj), style1)
                worksheet.write(2, 1, int(hx) / 2, style1)
                worksheet.write(2, 2, int(lhx) / 2, style1)
                worksheet.write(2, 3, int(hj) / 2, style1)
                worksheet.write(2, 4, Formula('D3-B3'), style1)
                worksheet.write(2, 5, Formula('D3-C3'), style1)
                worksheet.write(2, 6, Formula('D3-B3-C3'), style1)
                worksheet.write(3, 1, len(hxlist), style1)
                worksheet.write(3, 2, len(hxlist) + len(lhxlist), style1)
                worksheet.write(3, 3, len(hxlist) + len(lhxlist), style1)
                worksheet.write(3, 4, 26 - len(hxlist), style1)
                worksheet.write(3, 5, 26 - len(hxlist), style1)
                worksheet.write(4, 1, Formula('B3/B4'), style1)
                worksheet.write(4, 2, Formula('C3/C4'), style1)
                worksheet.write(4, 3, Formula('D3/D4'), style1)
                worksheet.write(4, 4, Formula('E3/E4'), style1)
                worksheet.write(4, 5, Formula('F3/F4'), style1)

    if os.path.exists(flow_path):
        book2.save(flow_path2)  # 保存新的excel，保存excel必须使用后缀名是.xls的，不是能是.xlsx的
        os.remove(flow_path)
        os.rename(flow_path2, flow_path)
    else:
        book2.save(flow_path)


def report_down_load(request):
    if request.method == 'POST':
        filename = request.POST.get('filename')

        path = os.path.abspath('excel/cmnetwin')
        if not (os.path.join(path, 'CMNET网工程计算.xls')):
            logger.info('没有可导出文件！')
            return JsonResponse({'status': 'err', 'msg': '没有可导出文件！！'})

        wb = open_exce(os.path.join(path, 'CMNET网工程计算.xls'))
        wb2 = copy(wb)

        if not filename:
            pro_name = ProjectInfo.objects.filter(id=2).values('name')
            pro_name = list(pro_name)[0]['name']

            response = HttpResponse(content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename="{0}"'.format(
                escape_uri_path(pro_name + '中继仿真.xls'))
            wb2.save(response)
            return response
        else:

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=' + filename + '.xls'
            wb2.save(response)
            return response
